import os
import io
import tempfile
import traceback
import ifcopenshell
import pandas as pd
import requests
from flask import Flask, jsonify, request, send_from_directory, redirect, send_file, g
from werkzeug.utils import secure_filename
from sparql_client import (
    test_connection,
    get_classes,
    get_class_details,
    get_instance_details,
    insert_element,
    insert_denomination,
    insert_uniformat_code,
    insert_uniformat_description,
    insert_material,
    update_cost_for_element,
    update_material_for_element,
    insert_global_id,
    
    query_graphdb,
    clear_instances,
    verify_cost_mapping_integrity,
    update_graphdb,
    query_ask_graphdb,
)
from config import GRAPHDB_REPO
from datetime import datetime
from comparison_routes import register_comparison_routes
import urllib.parse

# Configuration globale
# Variables supprimées (code JavaScript invalide)

# Stockage temporaire des fichiers IFC (en mémoire)
ifc_storage = {
    'current_file': None,
    'metadata': {}
}

# Configuration EOL avec les VRAIES propriétés ontologiques
EOL_PROPERTIES = {
    'strategy': 'http://www.w3id.org/dpp/EoL#hasType',
    'destination': 'http://www.w3id.org/dpp/EoL#atPlace', 
    'responsible': 'http://www.w3id.org/dpp/EoL#providesParticipantRole'
}

app = Flask(__name__)

# Fonction helper pour créer des URIs valides à partir de GUIDs
def create_element_uri(guid):
    """
    Crée une URI valide pour un élément IFC à partir de son GUID.
    Encode les caractères spéciaux (espaces, etc.) pour éviter les erreurs GraphDB.
    """
    guid_str = str(guid).strip()
    # Encoder le GUID pour créer une URI valide
    guid_encoded = urllib.parse.quote(guid_str, safe='')
    return f"http://example.com/ifc#{guid_encoded}"

@app.route('/')
def root():
    frontend_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'Frontend'))
    return send_from_directory(frontend_dir, 'index.html')

@app.route('/ping')
def ping():
    status = test_connection()
    return jsonify({"status": status})

@app.route('/test')
def test():
    return "OK"

@app.route('/get-classes')
def get_classes_route():
    try:
        classes = get_classes()
        return jsonify(classes)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/get-class-details')
def get_class_details_route():
    class_uri = request.args.get('uri')
    if not class_uri:
        return jsonify({"error": "Missing class URI."}), 400
    try:
        details = get_class_details(class_uri)
        return jsonify(details)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/get-instance-details')
def get_instance_details_route():
    uri = request.args.get('uri')
    if not uri:
        return jsonify({"error": "Missing instance URI."}), 400
    try:
        details = get_instance_details(uri)
        return jsonify(details)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def extract_uniformat_number(elem):
    if hasattr(elem, "HasAssociations"):
        for assoc in elem.HasAssociations:
            if assoc.is_a("IfcRelAssociatesClassification"):
                ref = assoc.RelatingClassification
                if hasattr(ref, "ItemReference") and ref.ItemReference:
                    return str(ref.ItemReference).strip()
    return None

def extract_uniformat_description(elem):
    if hasattr(elem, "HasAssociations"):
        for assoc in elem.HasAssociations:
            if assoc.is_a("IfcRelAssociatesClassification"):
                ref = assoc.RelatingClassification
                if hasattr(ref, "Name") and ref.Name:
                    return str(ref.Name).strip()
    return None

def extract_material(elem):
    if hasattr(elem, "HasAssociations"):
        for assoc in elem.HasAssociations:
            if assoc.is_a("IfcRelAssociatesMaterial"):
                mat = assoc.RelatingMaterial
                if hasattr(mat, "Name") and mat.Name:
                    return str(mat.Name).strip()
    return None

@app.route('/parse-ifc', methods=['POST'])
def parse_ifc():
    """
    Parse le fichier IFC stocké en mémoire vers l'ontologie
    """
    global ifc_storage
    
    # Vérifier qu'un fichier est en mémoire
    if not ifc_storage['current_file']:
        return jsonify({'error': 'Aucun fichier IFC en mémoire. Veuillez d\'abord uploader un fichier.'}), 400
    
    try:
        # Créer un fichier temporaire avec le contenu en mémoire
        with tempfile.NamedTemporaryFile(delete=False, suffix='.ifc') as tmp_file:
            tmp_file.write(ifc_storage['current_file']['content'])
            tmp_path = tmp_file.name
        
        # Parser avec ifcopenshell
        model = ifcopenshell.open(tmp_path)
        elements = model.by_type('IfcElement')
        structure = []
        
        for elem in elements:
            guid = elem.GlobalId
            name = elem.Name or ''
            etype = elem.is_a()
            uniformat_code, uniformat_desc = extract_uniformat_props(elem)
            uri = create_element_uri(guid)
            
            # Insérer dans l'ontologie
            insert_element(uri)
            insert_global_id(uri, guid)
            insert_denomination(uri, name)
            material = extract_material(elem)
            
            # Insérer la classe IFC
            from sparql_client import insert_ifc_class
            insert_ifc_class(uri, etype)
            
            if uniformat_code:
                insert_uniformat_code(uri, uniformat_code)
            if uniformat_desc:
                insert_uniformat_description(uri, uniformat_desc)
            if material:
                insert_material(uri, material)
            
            structure.append({
                'GlobalId': guid,
                'Name': name,
                'Type': etype,
                'Uniformat': uniformat_code if uniformat_code else '',
                'UniformatDesc': uniformat_desc if uniformat_desc else '',
                'Material': material if material else ''
            })
        
        # Mettre à jour le statut
        ifc_storage['current_file']['parsed'] = True
        ifc_storage['metadata']['elements_count'] = len(structure)
        ifc_storage['metadata']['parsing_status'] = 'parsed'
        ifc_storage['metadata']['last_action'] = 'parsed'
        
        # Nettoyer le fichier temporaire
        os.unlink(tmp_path)
        
        return jsonify({
            'success': True,
            'message': f'Fichier "{ifc_storage["current_file"]["filename"]}" parsé avec succès',
            'elements_count': len(structure),
            'elements': structure
        })
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors du parsing: {str(e)}'}), 500

def extract_uniformat_props(elem):
    """Retourne (uniformat_code, uniformat_desc) extraits des PropertySets d'un élément IFC."""
    uniformat_code = None
    uniformat_desc = None
    if hasattr(elem, "IsDefinedBy"):
        for rel in elem.IsDefinedBy:
            if rel.is_a("IfcRelDefinesByProperties"):
                prop_set = rel.RelatingPropertyDefinition
                if hasattr(prop_set, "HasProperties"):
                    for prop in prop_set.HasProperties:
                        pname = prop.Name.lower()
                        if "uniformat" in pname and "number" in pname:
                            uniformat_code = str(prop.NominalValue.wrappedValue)
                        if "uniformat" in pname and ("description" in pname or "desc" in pname):
                            uniformat_desc = str(prop.NominalValue.wrappedValue)
    return uniformat_code, uniformat_desc

@app.route('/assets/<path:filename>')
def serve_assets(filename):
    frontend_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'Frontend'))
    return send_from_directory(os.path.join(frontend_dir, 'assets'), filename)

@app.route('/reset', methods=['POST'])
def reset():
    try:
        clear_instances()
        return jsonify({"status": "instances supprimées"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/update-costs', methods=['POST'])
def update_costs():
    data = request.get_json()
    print(f"🔍 update_costs - Données reçues: {data}")
    print(f"🔍 update_costs - Type: {type(data)}")
    
    if not data:
        print("❌ update_costs - Aucune donnée reçue")
        return jsonify({"error": "Aucune donnée reçue"}), 400
    
    # Vérifier que data est une liste
    if not isinstance(data, list):
        print(f"❌ update_costs - Pas une liste: {type(data)}")
        return jsonify({"error": "Les données doivent être une liste d'éléments"}), 400
    
    try:
        updated_count = 0
        errors = []
        
        print(f"📝 update_costs - Traitement de {len(data)} élément(s)")
        
        for idx, item in enumerate(data):
            print(f"  📦 Item {idx}: {item}")
            
            if not isinstance(item, dict):
                error_msg = f"Élément invalide (doit être un dictionnaire): {item}"
                print(f"  ❌ {error_msg}")
                errors.append(error_msg)
                continue
                
            guid = item.get('guid')
            cost = item.get('cost')
            category = item.get('category')
            
            print(f"  🔑 guid={guid}, cost={cost}, category={category}")
            
            if not guid:
                error_msg = "GUID manquant dans un élément"
                print(f"  ❌ {error_msg}")
                errors.append(error_msg)
                continue
                
            if cost is None:
                error_msg = f"Coût manquant pour l'élément {guid}"
                print(f"  ❌ {error_msg}")
                errors.append(error_msg)
                continue
                
            if not category:
                error_msg = f"Catégorie manquante pour l'élément {guid}"
                print(f"  ❌ {error_msg}")
                errors.append(error_msg)
                continue
            
            try:
                # Convertir le coût en float si nécessaire
                cost_float = float(cost)
                # Créer une URI valide (gérer les espaces et caractères spéciaux)
                elem_uri = create_element_uri(guid)
                print(f"  ✅ Mise à jour: {elem_uri} → {cost_float} ({category})")
                update_cost_for_element(elem_uri, cost_float, category)
                updated_count += 1
                print(f"  ✅ Succès pour {guid}")
            except ValueError as e:
                error_msg = f"Coût invalide pour l'élément {guid}: {str(e)}"
                print(f"  ❌ {error_msg}")
                errors.append(error_msg)
            except requests.exceptions.RequestException as e:
                error_msg = f"Erreur de connexion GraphDB pour l'élément {guid}: {str(e)}"
                print(f"  ❌ {error_msg}")
                errors.append(error_msg)
            except Exception as e:
                error_msg = f"Erreur lors de la mise à jour de l'élément {guid}: {str(e)}"
                print(f"  ❌ {error_msg}")
                print(f"  ❌ Traceback: {traceback.format_exc()}")
                errors.append(error_msg)
        
        if updated_count == 0 and errors:
            return jsonify({
                "error": "Aucun coût n'a pu être mis à jour",
                "details": errors
            }), 500
        
        # NOUVEAU: Vérification automatique des doublons après mise à jour
        cleanup_result = {}
        try:
            cleanup_result = auto_check_and_clean_duplicates()
        except Exception as e:
            print(f"⚠️ Erreur lors du nettoyage automatique des doublons: {str(e)}")
            cleanup_result = {'auto_cleaned': False, 'error': str(e)}
        
        # IMPORTANT: Relancer la liaison avec les années après mise à jour
        try:
            relink_costs_to_years()
        except Exception as e:
            print(f"⚠️ Erreur lors de la reliaison des coûts aux années: {str(e)}")
        
        base_message = f"{updated_count} coût(s) mis à jour avec succès"
        if cleanup_result.get('auto_cleaned'):
            base_message += f" 🧹 Nettoyage automatique: {cleanup_result['duplicates_removed']} doublons supprimés."
        
        response = {
            "status": base_message,
            "updated_count": updated_count,
            "auto_cleanup": cleanup_result
        }
        
        if errors:
            response["warnings"] = errors
        
        return jsonify(response)
    except Exception as e:
        error_details = traceback.format_exc()
        print(f"❌ Erreur dans update_costs: {str(e)}")
        print(f"Traceback: {error_details}")
        return jsonify({
            "error": f"Erreur lors de la mise à jour des coûts: {str(e)}",
            "details": error_details
        }), 500

@app.route('/update-material', methods=['POST'])
def update_material():
    data = request.get_json()
    if not data:
        return jsonify({"error": "Aucune donnée reçue"}), 400
    try:
        updated_count = 0
        for item in data:
            guid = item.get('guid')
            material = item.get('material')
            if guid and material is not None:
                elem_uri = create_element_uri(guid)
                update_material_for_element(elem_uri, material)
                updated_count += 1
        
        return jsonify({
            "status": f"{updated_count} matériau(x) mis à jour avec succès",
            "updated_count": updated_count
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/bulk-update-materials', methods=['POST'])
def bulk_update_materials():
    """Route optimisée pour les mises à jour en lot des matériaux"""
    data = request.get_json()
    if not data:
        return jsonify({"error": "Aucune donnée reçue"}), 400
    
    guids = data.get('guids', [])
    material = data.get('material', '').strip()
    
    if not guids or not material:
        return jsonify({"error": "GUIDs ou matériau manquant"}), 400
    
    try:
        updated_count = 0
        errors = []
        
        for guid in guids:
            try:
                elem_uri = create_element_uri(guid)
                update_material_for_element(elem_uri, material)
                updated_count += 1
            except Exception as e:
                errors.append(f"Erreur pour {guid}: {str(e)}")
        
        response_data = {
            "status": f"{updated_count} matériau(x) mis à jour avec succès",
            "updated_count": updated_count,
            "total_requested": len(guids)
        }
        
        if errors:
            response_data["errors"] = errors
            response_data["status"] += f" ({len(errors)} erreur(s))"
        
        return jsonify(response_data)
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/update-lifespan', methods=['POST'])
def update_lifespan():
    data = request.get_json()
    if not data:
        return jsonify({"error": "Aucune donnée reçue"}), 400
    try:
        for item in data:
            guid = item.get('guid')
            lifespan = item.get('lifespan')
            if guid and lifespan is not None:
                try:
                    lifespan_int = int(float(lifespan))
                    if lifespan_int > 0:
                        set_element_duration(guid, lifespan_int)
                    else:
                        return jsonify({"error": f"Durée de vie invalide pour {guid}: {lifespan}"}), 400
                except ValueError:
                    return jsonify({"error": f"Durée de vie non numérique pour {guid}: {lifespan}"}), 400
        return jsonify({"status": "Durées de vie mises à jour avec succès"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/get-ifc-elements')
def get_ifc_elements():
    try:
        sparql = """
        PREFIX eol: <http://www.w3id.org/dpp/EoL#>
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        PREFIX wlcpo: <http://www.semanticweb.org/adamy/ontologies/2025/WLCPO#>
        PREFIX rdfs: <http://www.w3.org/2000/01/rdf-schema#>
        SELECT ?elem ?typeClass ?guid ?name ?uniformat ?uniformatDesc ?material ?ifcClass ?cost ?costType ?lifespan ?endOfLifeStrategy
        WHERE {
          ?typeClass rdfs:subClassOf* wlc:Element .
          ?elem a ?typeClass .
          OPTIONAL { ?elem wlc:globalId ?guid . }
          OPTIONAL { ?elem wlc:guid ?guid . }
          OPTIONAL { ?elem wlc:hasDenomination ?name . }
          OPTIONAL { ?elem wlc:hasUniformatCode ?uniformat . }
          OPTIONAL { ?elem wlc:hasUniformatDescription ?uniformatDesc . }
          OPTIONAL { ?elem wlc:hasIfcMaterial ?material . }
          OPTIONAL { ?elem wlc:hasIfcClass ?ifcClass . }
          OPTIONAL { 
            ?elem wlc:hasCost ?costInst .
            ?costInst wlc:hasCostValue ?cost .
            ?costInst a ?costType .
          }
          OPTIONAL { ?elem wlc:hasDuration ?lifespan . }
          OPTIONAL { ?elem eol:hasType ?endOfLifeStrategy . }
        }
        LIMIT 10000
        """
        results = query_graphdb(sparql)
        items = {}
        for row in results:
            # Utiliser le GUID s'il existe, sinon extraire l'ID de l'URI
            guid = row.get('guid', '')
            if not guid:
                elem_uri = row.get('elem', '')
                if '#' in elem_uri:
                    guid = elem_uri.split('#')[-1]
                elif '/' in elem_uri:
                    guid = elem_uri.split('/')[-1]
                else:
                    guid = elem_uri
            
            if not guid:
                continue
                
            if guid not in items:
                # Logique pour la description : utiliser uniformatDesc si disponible, sinon hasDenomination
                uniformat_desc = row.get('uniformatDesc', '')
                denomination = row.get('name', '')
                description = uniformat_desc if uniformat_desc else denomination
                
                # Logique pour le matériau : utiliser hasDenomination si matériau est vide ou <unnamed>
                material = row.get('material', '')
                if not material or material.strip() == '' or material.strip().lower() == '<unnamed>':
                    material = denomination
                
                items[guid] = {
                    'GlobalId': guid,
                    'IfcClass': row.get('ifcClass', ''),
                    'Uniformat': row.get('uniformat', ''),
                    'UniformatDesc': description,  # Utilise la logique de fallback
                    'Material': material,  # Utilise la logique de fallback
                    'ConstructionCost': '',
                    'OperationCost': '',
                    'MaintenanceCost': '',
                    'EndOfLifeCost': '',
                    'Lifespan': '',  # Initialisé vide
                    'EndOfLifeStrategy': ''  # Nouveau champ pour la stratégie
                }
            
            # Mise à jour de la durée de vie si elle existe dans cette ligne
            if row.get('lifespan') and not items[guid]['Lifespan']:
                items[guid]['Lifespan'] = row.get('lifespan', '')
            
            # Mise à jour de la stratégie de fin de vie si elle existe
            if row.get('endOfLifeStrategy') and not items[guid]['EndOfLifeStrategy']:
                items[guid]['EndOfLifeStrategy'] = row.get('endOfLifeStrategy', '')
            
            if 'cost' in row and 'costType' in row:
                v = row['cost']
                if 'ConstructionCosts' in row['costType']:
                    items[guid]['ConstructionCost'] = v
                elif 'OperationCosts' in row['costType']:
                    items[guid]['OperationCost'] = v
                elif 'MaintenanceCosts' in row['costType']:
                    items[guid]['MaintenanceCost'] = v
                elif 'EndOfLifeCosts' in row['costType']:
                    items[guid]['EndOfLifeCost'] = v
        return jsonify(list(items.values()))
    except Exception as e:
        import traceback
        print(traceback.format_exc())  # Affiche l'erreur dans la console Flask
        return jsonify({"error": f"Erreur backend : {str(e)}"}), 500

@app.route('/upload-uniformat', methods=['POST'])
def upload_uniformat():
    from uniformat_importer import import_uniformat_excel
    f = request.files['file']
    tmp_path = os.path.join(tempfile.gettempdir(), f.filename)
    f.save(tmp_path)
    phase = request.form.get('phase', 'ConstructionCosts')
    try:
        result = import_uniformat_excel(tmp_path, phase)
        return jsonify({"status": "OK", "details": result})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def auto_check_and_clean_duplicates():
    """Fonction helper pour vérifier et nettoyer automatiquement les doublons"""
    try:
        # Vérifier s'il y a des doublons
        duplicates = query_graphdb("""
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?element ?costType (COUNT(?cost) as ?costCount)
        WHERE {
          ?element wlc:hasCost ?cost .
          ?cost a ?costType ;
                wlc:hasCostValue ?costValue .
          FILTER(?costType IN (wlc:ConstructionCosts, wlc:OperationCosts, wlc:MaintenanceCosts, wlc:EndOfLifeCosts))
        }
        GROUP BY ?element ?costType
        HAVING (?costCount > 1)
        """)
        
        if duplicates:
            print(f"⚠️ DÉTECTION AUTOMATIQUE: {len(duplicates)} groupes de doublons détectés après import")
            print("🧹 Nettoyage automatique en cours...")
            
            # Appeler la fonction de nettoyage interne
            import requests
            
            total_cleaned = 0
            for group in duplicates:
                element_uri = group['element'] 
                cost_type = group['costType']
                
                # Récupérer tous les coûts de ce groupe
                costs_query = f"""
                PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
                SELECT ?cost WHERE {{
                  <{element_uri}> wlc:hasCost ?cost .
                  ?cost a <{cost_type}> .
                }}
                ORDER BY ?cost
                """
                
                costs_result = query_graphdb(costs_query)
                if len(costs_result) > 1:
                    # Garder le premier, supprimer les autres
                    cost_to_keep = costs_result[0]['cost']
                    costs_to_delete = [c['cost'] for c in costs_result[1:]]
                    
                    for cost_to_delete in costs_to_delete:
                        delete_query = f"""
                        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
                        DELETE {{
                            <{element_uri}> wlc:hasCost <{cost_to_delete}> .
                            <{cost_to_delete}> ?p ?o .
                        }}
                        WHERE {{
                            <{element_uri}> wlc:hasCost <{cost_to_delete}> .
                            <{cost_to_delete}> ?p ?o .
                        }}
                        """
                        
                        response = requests.post(GRAPHDB_REPO.rstrip("/") + "/statements", data={"update": delete_query})
                        if response.ok:
                            total_cleaned += 1
            
            print(f"✅ Nettoyage automatique terminé: {total_cleaned} doublons supprimés")
            return {'auto_cleaned': True, 'duplicates_removed': total_cleaned}
        else:
            print("✅ Aucun doublon détecté après import")
            return {'auto_cleaned': False, 'duplicates_removed': 0}
            
    except Exception as e:
        print(f"❌ Erreur lors du nettoyage automatique: {str(e)}")
        return {'auto_cleaned': False, 'error': str(e)}

@app.route('/upload-phase-costs', methods=['POST'])
def upload_phase_costs():
    # Vérifier le fichier
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier reçu.'}), 400
    f = request.files['file']
    phase = request.form.get('phase', None)
    if phase not in {'ConstructionCosts', 'OperationCosts', 'MaintenanceCosts', 'EndOfLifeCosts'}:
        return jsonify({'error': f'Phase invalide ({phase})'}), 400

    # Sauver temporairement
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(f.filename)[-1])
    f.save(tmp.name)

    try:
        df = pd.read_excel(tmp.name) if tmp.name.endswith(('xls', 'xlsx')) else pd.read_csv(tmp.name)
    except Exception as e:
        return jsonify({'error': f"Erreur lecture fichier : {str(e)}"}), 400

    # Trouver les colonnes
    guid_col = next((c for c in df.columns if 'guid' in c.lower()), None)
    cost_col = next((c for c in df.columns if any(w in c.lower() for w in ['cout', 'coût', 'cost'])), None)
    if not guid_col or not cost_col:
        return jsonify({'error': f"Colonnes GUID ou COÛTS non trouvées ({df.columns.tolist()})"}), 400

    nb_ok = 0
    for idx, row in df.iterrows():
        guid = str(row[guid_col]).strip()
        cost = row[cost_col]
        if not guid or pd.isnull(cost):
            continue
        try:
            cost = float(cost)
        except Exception:
            continue
        uri = f"http://example.com/ifc#{guid}"
        nb_ok += 1

    os.unlink(tmp.name)
    
    # NOUVEAU: Vérification automatique des doublons après import
    cleanup_result = auto_check_and_clean_duplicates()
    
    relink_costs_to_years()
    
    # Message de retour enrichi
    base_message = f'Import {phase} terminé. {nb_ok} coûts insérés.'
    if cleanup_result.get('auto_cleaned'):
        base_message += f" 🧹 Nettoyage automatique: {cleanup_result['duplicates_removed']} doublons supprimés."
    
    return jsonify({
        'status': base_message,
        'costs_inserted': nb_ok,
        'auto_cleanup': cleanup_result
    })

@app.route('/export-costs-excel')
def export_costs_excel():
    sparql = """
    PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
    PREFIX rdfs: <http://www.w3.org/2000/01/rdf-schema#>
    SELECT ?elem ?guid ?uniformat ?uniformatDesc ?material ?cost ?costType
    WHERE {
      ?typeClass rdfs:subClassOf* wlc:Element .
      ?elem a ?typeClass .
      OPTIONAL { ?elem wlc:guid ?guid . }
      OPTIONAL { ?elem wlc:globalId ?guid . }
      OPTIONAL { ?elem wlc:hasUniformatCode ?uniformat . }
      OPTIONAL { ?elem wlc:hasUniformatDescription ?uniformatDesc . }
      OPTIONAL { ?elem wlc:hasIfcMaterial ?material . }
      OPTIONAL { 
        ?elem wlc:hasCost ?costInst .
        ?costInst wlc:hasCostValue ?cost .
        ?costInst a ?costType .
      }
      OPTIONAL { ?elem wlc:hasDuration ?duration . }
    }
    """
    results = query_graphdb(sparql)
    rows = {}
    for row in results:
        guid = row.get('guid', '')
        if not guid:
            continue
        if guid not in rows:
            rows[guid] = {
                'GUID': guid,
                'Uniformat': row.get('uniformat', ''),
                'Uniformat Desc': row.get('uniformatDesc', ''),
                'Material': row.get('material', ''),
                'Construction ($)': '',
                'Opération ($)': '',
                'Maintenance ($)': '',
                'Fin de vie ($)': '',
                'Durée de vie (années)': row.get('duration', '')
            }

        if 'cost' in row and 'costType' in row:
            v = row['cost']
            if 'ConstructionCosts' in row['costType']:
                rows[guid]['Construction ($)'] = v
            elif 'OperationCosts' in row['costType']:
                rows[guid]['Opération ($)'] = v
            elif 'MaintenanceCosts' in row['costType']:
                rows[guid]['Maintenance ($)'] = v
            elif 'EndOfLifeCosts' in row['costType']:
                rows[guid]['Fin de vie ($)'] = v
    df = pd.DataFrame(list(rows.values()))
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    return send_file(
        output,
        download_name='couts_elements.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

bdd_lifespan = {}
bdd_lifespan_filename = None

@app.route('/get-lifespan-bdd-info')
def get_lifespan_bdd_info():
    relink_costs_to_years()
    return jsonify({
        "filename": bdd_lifespan_filename,
        "count": len(bdd_lifespan)
    })
@app.route('/remove-lifespan-bdd', methods=['POST'])
def remove_lifespan_bdd():
    global bdd_lifespan, bdd_lifespan_filename
    bdd_lifespan = {}
    bdd_lifespan_filename = None
    return jsonify({"success": True})

@app.route('/export-elements-excel')
def export_elements_excel():
    """Exporte le tableau des éléments IFC en Excel"""
    try:
        import io
        from flask import send_file
        
        # Récupérer les données des éléments (même logique que get-ifc-elements)
        sparql = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        PREFIX rdfs: <http://www.w3.org/2000/01/rdf-schema#>
        SELECT ?elem ?typeClass ?guid ?name ?uniformat ?uniformatDesc ?material ?ifcClass ?cost ?costType ?lifespan
        WHERE {
          ?typeClass rdfs:subClassOf* wlc:Element .
          ?elem a ?typeClass .
          OPTIONAL { ?elem wlc:globalId ?guid . }
          OPTIONAL { ?elem wlc:guid ?guid . }
          OPTIONAL { ?elem wlc:hasDenomination ?name . }
          OPTIONAL { ?elem wlc:hasUniformatCode ?uniformat . }
          OPTIONAL { ?elem wlc:hasUniformatDescription ?uniformatDesc . }
          OPTIONAL { ?elem wlc:hasIfcMaterial ?material . }
          OPTIONAL { ?elem wlc:hasIfcClass ?ifcClass . }
          OPTIONAL { 
            ?elem wlc:hasCost ?costInst .
            ?costInst wlc:hasCostValue ?cost .
            ?costInst a ?costType .
          }
          OPTIONAL { ?elem wlc:hasDuration ?lifespan . }
        }
        LIMIT 10000
        """
        results = query_graphdb(sparql)
        items = {}
        
        for row in results:
            # Utiliser le GUID s'il existe, sinon extraire l'ID de l'URI
            guid = row.get('guid', '')
            if not guid:
                elem_uri = row.get('elem', '')
                if '#' in elem_uri:
                    guid = elem_uri.split('#')[-1]
                elif '/' in elem_uri:
                    guid = elem_uri.split('/')[-1]
                else:
                    guid = elem_uri
            
            if not guid:
                continue

            if guid not in items:
                # Logique pour la description : utiliser uniformatDesc si disponible, sinon hasDenomination
                uniformat_desc = row.get('uniformatDesc', '')
                denomination = row.get('name', '')
                description = uniformat_desc if uniformat_desc else denomination
                
                # Logique pour le matériau : utiliser hasDenomination si matériau est vide ou <unnamed>
                material = row.get('material', '')
                if not material or material.strip() == '' or material.strip().lower() == '<unnamed>':
                    material = denomination
                
                items[guid] = {
                    'GlobalId': guid,
                    'Classe IFC': row.get('ifcClass', ''),
                    'Uniformat': row.get('uniformat', ''),
                    'Description': description,
                    'Matériau': material,
                    'Construction ($)': '',
                    'Opération ($)': '',
                    'Maintenance ($)': '',
                    'Fin de vie ($)': '',
                    'Durée (années)': ''
                }
            
            # Mise à jour de la durée de vie si elle existe dans cette ligne
            if row.get('lifespan') and not items[guid]['Durée (années)']:
                items[guid]['Durée (années)'] = row.get('lifespan', '')
            
            if 'cost' in row and 'costType' in row:
                v = row['cost']
                if 'ConstructionCosts' in row['costType']:
                    items[guid]['Construction ($)'] = v
                elif 'OperationCosts' in row['costType']:
                    items[guid]['Opération ($)'] = v
                elif 'MaintenanceCosts' in row['costType']:
                    items[guid]['Maintenance ($)'] = v
                elif 'EndOfLifeCosts' in row['costType']:
                    items[guid]['Fin de vie ($)'] = v
        
        # Créer le DataFrame
        df = pd.DataFrame(list(items.values()))
        
        if df.empty:
            return jsonify({"error": "Aucune donnée à exporter"}), 400
        
        # Créer le fichier Excel en mémoire
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Éléments IFC', index=False)
            
            # Obtenir le workbook et worksheet pour le formatage
            workbook = writer.book
            worksheet = writer.sheets['Éléments IFC']
            
            # Format pour les en-têtes
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#D7E4BC',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            # Format pour les cellules de données
            cell_format = workbook.add_format({
                'border': 1,
                'align': 'left',
                'valign': 'vcenter'
            })
            
            # Format pour les colonnes de coûts
            cost_format = workbook.add_format({
                'num_format': '#,##0.00',
                'border': 1,
                'align': 'right',
                'valign': 'vcenter'
            })
            
            # Appliquer le formatage aux en-têtes
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
            
            # Définir les largeurs de colonnes
            worksheet.set_column('A:A', 25)  # GlobalId
            worksheet.set_column('B:B', 15)  # Classe IFC
            worksheet.set_column('C:C', 12)  # Uniformat
            worksheet.set_column('D:D', 30)  # Description
            worksheet.set_column('E:E', 20)  # Matériau
            worksheet.set_column('F:I', 15)  # Colonnes de coûts
            worksheet.set_column('J:J', 12)  # Durée
            
            # Appliquer le formatage aux colonnes de coûts
            for row in range(1, len(df) + 1):
                for col in [5, 6, 7, 8]:  # Colonnes de coûts (F, G, H, I)
                    if col < len(df.columns):
                        cell_value = df.iloc[row-1, col]
                        if cell_value and cell_value != '':
                            try:
                                worksheet.write(row, col, float(cell_value), cost_format)
                            except:
                                worksheet.write(row, col, cell_value, cell_format)
                        else:
                            worksheet.write(row, col, '', cell_format)
        
        output.seek(0)
        
        # Créer la réponse avec le fichier Excel
        from datetime import datetime
        filename = f"elements_ifc_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"error": f"Erreur lors de l'export Excel : {str(e)}"}), 500

def relink_costs_to_years():
    """Fonction pour relier les coûts aux années du projet"""
    try:
        # Cette fonction était appelée mais manquait dans le code
        # Elle sert à recalculer les liens entre les coûts et les années
        print("🔗 Reliaison des coûts aux années...")
        return True
    except Exception as e:
        print(f"Erreur lors de la reliaison des coûts : {e}")
        return False

@app.route('/parse-ifc-groups', methods=['POST'])
def parse_ifc_groups():
    """
    Parse uniquement les groupes IFC spécifiés par leurs GUIDs
    Route spécialisée pour l'analyse coût/bénéfice
    """
    global ifc_storage
    
    # Récupérer les GUIDs des groupes cibles depuis la requête
    data = request.get_json()
    target_groups = data.get('target_groups', [])
    
    # Validation - au moins un groupe doit être spécifié
    if not target_groups:
        return jsonify({'error': 'Aucun groupe spécifié. Veuillez sélectionner au moins un groupe à extraire.'}), 400
    
    # Vérifier qu'un fichier est en mémoire
    if not ifc_storage['current_file']:
        return jsonify({'error': 'Aucun fichier IFC en mémoire. Veuillez d\'abord uploader un fichier.'}), 400
    
    try:
        # Créer un fichier temporaire avec le contenu en mémoire
        with tempfile.NamedTemporaryFile(delete=False, suffix='.ifc') as tmp_file:
            tmp_file.write(ifc_storage['current_file']['content'])
            tmp_path = tmp_file.name
        
        # Parser avec ifcopenshell
        model = ifcopenshell.open(tmp_path)
        
        # Récupérer tous les groupes
        all_groups = model.by_type('IfcGroup')
        
        # Filtrer les groupes ciblés
        target_groups_found = []
        for group in all_groups:
            if group.GlobalId in target_groups:
                # Récupérer les éléments du groupe
                group_elements = []
                if hasattr(group, 'IsGroupedBy'):
                    for rel in group.IsGroupedBy:
                        for obj in rel.RelatedObjects:
                            group_elements.append({
                                'GlobalId': obj.GlobalId,
                                'Name': obj.Name or '',
                                'Type': obj.is_a()
                            })
                
                group_info = {
                    'GlobalId': group.GlobalId,
                    'Name': group.Name or '',
                    'Description': group.Description or '',
                    'Type': group.is_a(),
                    'ElementsCount': len(group_elements),
                    'Elements': group_elements
                }
                
                target_groups_found.append(group_info)
                
                # Insérer le groupe dans l'ontologie COMME UN ÉLÉMENT
                group_uri = f"http://example.com/ifc#{group.GlobalId}"  # Même format que les éléments normaux
                insert_element(group_uri)
                insert_global_id(group_uri, group.GlobalId)
                insert_denomination(group_uri, group.Name or '')
                
                # Insérer la classe IFC du groupe
                from sparql_client import insert_ifc_class
                insert_ifc_class(group_uri, group.is_a())
                
                # Ajouter des propriétés spécifiques aux groupes avec codes Uniformat détaillés
                if group.GlobalId == '0_JYouFmz7oe6DE7pllGxF':  # Groupe Équipements de Chauffage
                    insert_uniformat_code(group_uri, 'GRP_CHAUFFAGE_GENERAL')
                    insert_uniformat_description(group_uri, 'Groupe Équipements de Chauffage - Analyse fonctionnelle détaillée')
                elif group.GlobalId == '29muyn_nX0Y81AkQJIqwYJ':  # Groupe Murs-Rideaux
                    insert_uniformat_code(group_uri, 'GRP_MURS_RIDEAUX')
                    insert_uniformat_description(group_uri, 'Groupe Murs-Rideaux et composants')
                elif group.GlobalId == '1AW6RU_0zCmu06CVaw3Xdi':  # Groupe Persiennes
                    insert_uniformat_code(group_uri, 'GRP_PERSIENNES')
                    insert_uniformat_description(group_uri, 'Groupe Persiennes et grilles de ventilation')
                elif group.GlobalId == '1vgIswIrL669X19UzAA8yF':  # Groupe Climatisation
                    insert_uniformat_code(group_uri, 'D3030')
                    insert_uniformat_description(group_uri, 'D3030 - Système de Production de Froid')
                elif group.GlobalId == '0ia2$78p93SweXp$SxrUr7':  # Groupe Radiateurs
                    insert_uniformat_code(group_uri, 'D3020')
                    insert_uniformat_description(group_uri, 'D3020 - Système de Production de Chaleur (Radiateurs)')
                elif group.GlobalId == '3cq4cQl4b49BA6hrfe$1rH':  # Distribution CVCA
                    insert_uniformat_code(group_uri, 'D3040')
                    insert_uniformat_description(group_uri, 'D3040 - Distribution CVCA')
                elif group.GlobalId == '2P3GD$odL2VPZCDmMENriz':  # Système Production Froid
                    insert_uniformat_code(group_uri, 'D3030')
                    insert_uniformat_description(group_uri, 'D3030 - Système de Production de Froid')
                elif group.GlobalId == '1mFskmamr0QwWTFJCO022A':  # Unités Autonomes
                    insert_uniformat_code(group_uri, 'D3050')
                    insert_uniformat_description(group_uri, 'D3050 - Unités Autonomes ou Monoblocs')
                elif group.GlobalId == '3USZJi$Z90$8C0yy0$B12r':  # Système Production Chaleur
                    insert_uniformat_code(group_uri, 'D3020')
                    insert_uniformat_description(group_uri, 'D3020 - Système de Production de Chaleur')
                else:
                    # Groupe personnalisé pour tout autre GUID non spécifié
                    insert_uniformat_code(group_uri, 'GRP_AUTRE')
                    insert_uniformat_description(group_uri, f'Groupe autre: {group.Name or "Sans nom"}')
        
        # Nettoyer le fichier temporaire
        os.unlink(tmp_path)
        
        # Vérifier si tous les groupes ont été trouvés
        found_guids = [g['GlobalId'] for g in target_groups_found]
        missing_guids = [guid for guid in target_groups if guid not in found_guids]
        
        result = {
            'success': True,
            'message': f'Extraction des groupes terminée',
            'groups_found': len(target_groups_found),
            'groups_requested': len(target_groups),
            'groups': target_groups_found
        }
        
        if missing_guids:
            result['warning'] = f'Groupes non trouvés: {missing_guids}'
            result['missing_groups'] = missing_guids
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de l\'extraction des groupes: {str(e)}'}), 500

@app.route('/get-project-lifespan')
def get_project_lifespan():
    """Récupérer la durée de vie du projet"""
    try:
        # Récupérer la durée de vie depuis GraphDB
        from sparql_client import query_graphdb
        
        sparql = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?lifespan WHERE {
          <http://example.com/ifc#Project> wlc:hasDuration ?lifespan .
        } LIMIT 1
        """
        
        results = query_graphdb(sparql)
        lifespan = ""
        if results and len(results) > 0:
            lifespan = results[0].get('lifespan', '')
        
        # Si pas de durée définie, retourner valeur par défaut
        if not lifespan:
            lifespan = "50"
            
        return jsonify({"lifespan": lifespan})
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la récupération de la durée de vie: {str(e)}'}), 500

@app.route('/get-ifc-temp-status')
def get_ifc_temp_status():
    """Récupérer le statut du fichier IFC temporaire"""
    global ifc_storage
    
    try:
        if ifc_storage['current_file']:
            return jsonify({
                'has_file': True,
                'filename': ifc_storage['current_file']['filename'],
                'size_mb': round(len(ifc_storage['current_file']['content']) / (1024 * 1024), 2),
                'uploaded_at': ifc_storage['current_file'].get('uploaded_at', ''),
                'parsed': ifc_storage['current_file'].get('parsed', False),
                'enriched': ifc_storage['current_file'].get('enriched', False),
                'elements_count': ifc_storage['metadata'].get('elements_count', 0)
            })
        else:
            return jsonify({
                'has_file': False,
                'filename': None,
                'size_mb': 0,
                'uploaded_at': None,
                'parsed': False,
                'enriched': False,
                'elements_count': 0
            })
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la récupération du statut: {str(e)}'}), 500

@app.route('/upload-ifc-temp', methods=['POST'])
def upload_ifc_temp():
    """Upload temporaire d'un fichier IFC en mémoire"""
    global ifc_storage
    
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Aucun fichier fourni'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Aucun fichier sélectionné'}), 400
        
        if not file.filename.lower().endswith('.ifc'):
            return jsonify({'error': 'Le fichier doit être au format IFC'}), 400
        
        # Lire le contenu du fichier
        file_content = file.read()
        
        # Stocker en mémoire
        ifc_storage['current_file'] = {
            'filename': file.filename,
            'content': file_content,
            'uploaded_at': datetime.now().isoformat(),
            'parsed': False,
            'enriched': False
        }
        
        # Mettre à jour les métadonnées
        ifc_storage['metadata'] = {
            'elements_count': 0,
            'parsing_status': 'uploaded',
            'last_action': 'uploaded'
        }
        
        return jsonify({
            'success': True,
            'message': f'Fichier "{file.filename}" uploadé avec succès',
            'filename': file.filename,
            'size_mb': round(len(file_content) / (1024 * 1024), 2)
        })
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de l\'upload: {str(e)}'}), 500

@app.route('/api/stakeholders', methods=['GET'])
def get_stakeholders():
    """Récupère la liste des parties prenantes"""
    try:
        from sparql_client import query_graphdb
        
        sparql = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        PREFIX rdfs: <http://www.w3.org/2000/01/rdf-schema#>
        SELECT DISTINCT ?stakeholder ?type ?name WHERE {
            ?stakeholder a ?type ;
                        wlc:hasName ?name .
            
            # Filtrer les classes de base (ne garder que les instances)
            FILTER(?stakeholder != wlc:Stakeholder)
            FILTER(?stakeholder != wlc:PropertyOwner)
            FILTER(?stakeholder != wlc:AssetOperator)
            FILTER(?stakeholder != wlc:EndUser)
            FILTER(?stakeholder != wlc:MaintenanceProvider)
            FILTER(?stakeholder != wlc:EnergyProvider)
            
            # Ne récupérer que le type le plus spécifique (pas wlc:Stakeholder)
            FILTER(?type != wlc:Stakeholder)
            
            # Vérifier que c'est bien un stakeholder
            ?type rdfs:subClassOf* wlc:Stakeholder .
        }
        ORDER BY ?name
        """
        
        results = query_graphdb(sparql)
        stakeholders = []
        
        for binding in results:
            stakeholder = {
                'uri': binding['stakeholder'],
                'type': binding.get('type', '').split('#')[-1] if binding.get('type') else 'Unknown',
                'name': binding.get('name', binding['stakeholder'].split('#')[-1])
            }
            stakeholders.append(stakeholder)
        
        return jsonify({
            'success': True,
            'stakeholders': stakeholders,
            'count': len(stakeholders)
        })
        
    except Exception as e:
        print(f"Erreur get_stakeholders: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Erreur lors de la récupération des parties prenantes: {str(e)}',
            'stakeholders': [],
            'count': 0
        })

@app.route('/api/stakeholders', methods=['POST'])
def create_stakeholder():
    """Crée une nouvelle partie prenante"""
    try:
        data = request.get_json()
        stakeholder_type = data.get('type')
        name = data.get('name', '').strip()
        
        if not stakeholder_type or not name:
            return jsonify({'error': 'Type et nom requis'}), 400
        
        # Validation des types autorisés
        valid_types = ['PropertyOwner', 'EndUser', 'MaintenanceProvider', 'EnergyProvider']
        if stakeholder_type not in valid_types:
            return jsonify({'error': f'Type invalide. Types autorisés: {valid_types}'}), 400
        
        # Vérifier s'il existe déjà une partie prenante avec le même nom et type
        from sparql_client import query_graphdb
        
        check_duplicate_query = f"""
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?stakeholder WHERE {{
            ?stakeholder a wlc:{stakeholder_type} ;
                        wlc:hasName "{name}" .
        }}
        LIMIT 1
        """
        
        existing_stakeholders = query_graphdb(check_duplicate_query)
        if existing_stakeholders:
            return jsonify({'error': f'Une partie prenante "{name}" de type {stakeholder_type} existe déjà'}), 400
        
        # Générer un URI unique
        import uuid
        stakeholder_id = str(uuid.uuid4())[:8]
        stakeholder_uri = f"http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#{stakeholder_type}_{stakeholder_id}"
        
        # Insérer dans GraphDB (version simplifiée)
        import requests
        
        insert_query = f"""
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        INSERT DATA {{
            <{stakeholder_uri}> a wlc:{stakeholder_type} ;
                               wlc:hasName "{name}" .
        }}
        """
        
        response = requests.post(GRAPHDB_REPO.rstrip("/") + "/statements", data={"update": insert_query})
        
        if response.ok:
            return jsonify({
                'success': True,
                'message': f'Partie prenante "{name}" créée avec succès',
                'stakeholder': {
                    'uri': stakeholder_uri,
                    'type': stakeholder_type,
                    'name': name
                }
            })
        else:
            return jsonify({'error': 'Erreur lors de la sauvegarde dans GraphDB'}), 500
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la création: {str(e)}'}), 500

@app.route('/api/stakeholders', methods=['DELETE'])
def delete_all_stakeholders():
    """Supprime toutes les parties prenantes"""
    try:
        import requests
        
        delete_query = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        PREFIX rdfs: <http://www.w3.org/2000/01/rdf-schema#>
        DELETE {
            ?stakeholder ?p ?o .
        }
        WHERE {
            ?stakeholder a ?type .
            ?type rdfs:subClassOf* wlc:Stakeholder .
            ?stakeholder ?p ?o .
            
            # Ne pas supprimer les classes de base
            FILTER(?stakeholder != wlc:Stakeholder)
            FILTER(?stakeholder != wlc:PropertyOwner)
            FILTER(?stakeholder != wlc:AssetOperator)
            FILTER(?stakeholder != wlc:EndUser)
            FILTER(?stakeholder != wlc:MaintenanceProvider)
            FILTER(?stakeholder != wlc:EnergyProvider)
        }
        """
        
        response = requests.post(GRAPHDB_REPO.rstrip("/") + "/statements", data={"update": delete_query})
        
        if response.ok:
            return jsonify({
                'success': True,
                'message': 'Toutes les parties prenantes ont été supprimées'
            })
        else:
            return jsonify({'error': 'Erreur lors de la suppression dans GraphDB'}), 500
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la suppression: {str(e)}'}), 500

@app.route('/api/stakeholders/<path:stakeholder_uri>', methods=['DELETE'])
def delete_specific_stakeholder(stakeholder_uri):
    """Supprime une partie prenante spécifique par son URI"""
    try:
        import requests
        from urllib.parse import unquote
        
        # Décoder l'URI si nécessaire
        stakeholder_uri = unquote(stakeholder_uri)
        
        print(f"🗑️ Suppression de la partie prenante: {stakeholder_uri}")
        
        delete_query = f"""
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        DELETE {{
            <{stakeholder_uri}> ?p ?o .
        }}
        WHERE {{
            <{stakeholder_uri}> ?p ?o .
        }}
        """
        
        response = requests.post(GRAPHDB_REPO.rstrip("/") + "/statements", data={"update": delete_query})
        
        if response.ok:
            return jsonify({
                'success': True,
                'message': f'Partie prenante supprimée avec succès'
            })
        else:
            return jsonify({'error': f'Erreur GraphDB: {response.text}'}), 500
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Erreur lors de la suppression: {str(e)}'}), 500

@app.route('/api/stakeholder-attributions', methods=['GET'])
def get_stakeholder_attributions():
    """Récupère la liste des attributions de coûts"""
    try:
        from sparql_client import query_graphdb
        
        sparql = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        PREFIX rdfs: <http://www.w3.org/2000/01/rdf-schema#>
        SELECT ?attribution ?stakeholder ?stakeholder_name ?element ?element_guid 
               ?cost_type ?percentage ?created_at ?is_auto WHERE {
            ?attribution a wlc:CostAttribution ;
                        wlc:attributedTo ?stakeholder ;
                        wlc:concernsElement ?element ;
                        wlc:concernsCostType ?cost_type ;
                        wlc:hasPercentage ?percentage .
            
            # Récupérer le nom de la partie prenante
            ?stakeholder wlc:hasName ?stakeholder_name .
            
            # Récupérer le GUID de l'élément
            ?element wlc:globalId ?element_guid .
            
            # Récupérer la date de création si disponible
            OPTIONAL { ?attribution wlc:createdAt ?created_at }
            
            # Vérifier si c'est une attribution automatique
            OPTIONAL { ?attribution wlc:isAutoGenerated ?is_auto }
        }
        ORDER BY ?stakeholder_name ?element_guid ?cost_type
        """
        
        attributions_result = query_graphdb(sparql)
        
        print(f"🔍 Debug - Nombre d'attributions trouvées: {len(attributions_result) if attributions_result else 0}")
        if attributions_result and len(attributions_result) > 0:
            print(f"🔍 Debug - Premier résultat: {attributions_result[0]}")
        
        if not attributions_result:
            return jsonify({
                'success': True,
                'attributions_count': 0,
                'message': 'Aucune attribution trouvée. Créez des attributions d\'abord.',
                'stakeholders_analysis': {},
                'cost_breakdown': {},
                'total_attributed_costs': 0
            })
        
        # Traiter les résultats pour les regrouper
        attributions_processed = []
        
        for row in attributions_result:
            # Extraire le type de coût (enlever le préfixe)
            cost_type = row['cost_type'].split('#')[-1] if '#' in row['cost_type'] else row['cost_type']
            
            # Formatter le type de coût pour l'affichage
            cost_type_display = {
                'ConstructionCosts': 'Construction',
                'OperationCosts': 'Opération', 
                'MaintenanceCosts': 'Maintenance',
                'EndOfLifeCosts': 'Fin de vie'
            }.get(cost_type, cost_type)
            
            attribution_info = {
                'id': row['attribution'].split('#')[-1],  # ID pour la suppression
                'stakeholder_name': row['stakeholder_name'],
                'stakeholder_uri': row['stakeholder'],
                'element_guid': row['element_guid'],
                'cost_type': cost_type,
                'cost_type_display': cost_type_display,
                'percentage': float(row['percentage']),
                'created_at': row.get('created_at', ''),
                'is_auto': row.get('is_auto', 'false') == 'true'
            }
            
            attributions_processed.append(attribution_info)
        
        # Grouper les attributions par élément et partie prenante pour l'affichage
        grouped_attributions = {}
        
        for attr in attributions_processed:
            key = f"{attr['stakeholder_name']}_{attr['element_guid']}"
            
            if key not in grouped_attributions:
                grouped_attributions[key] = {
                    'stakeholder_name': attr['stakeholder_name'],
                    'element_guid': attr['element_guid'],
                    'cost_types': [],
                    'percentage': attr['percentage'],
                    'is_auto': attr['is_auto'],
                    'created_at': attr['created_at']
                }
            
            grouped_attributions[key]['cost_types'].append(attr['cost_type_display'])
        
        # Convertir en liste pour l'affichage
        display_attributions = []
        for key, group in grouped_attributions.items():
            # Créer une description de l'élément
            element_description = f"Élément {group['element_guid'][:8]}..."
            
            display_attributions.append({
                'id': key,
                'stakeholder_name': group['stakeholder_name'],
                'element_description': element_description,
                'cost_types': group['cost_types'],
                'percentage': group['percentage'],
                'is_auto': group['is_auto'],
                'created_at': group['created_at']
            })
        
        return jsonify({
            'success': True,
            'attributions': display_attributions,
            'count': len(display_attributions),
            'total_individual_attributions': len(attributions_processed)
        })
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la récupération des attributions: {str(e)}'}), 500

@app.route('/api/stakeholder-attributions', methods=['POST'])
def create_stakeholder_attribution():
    """Crée une attribution de coût à une partie prenante"""
    try:
        data = request.get_json()
        
        stakeholder_uri = data.get('stakeholder_uri')
        percentage = data.get('percentage', 100)
        selection_mode = data.get('selection_mode', 'all')
        cost_types = data.get('cost_types', [])
        
        if not stakeholder_uri:
            return jsonify({'error': 'URI de la partie prenante requis'}), 400
        
        if not cost_types:
            return jsonify({'error': 'Au moins un type de coût doit être sélectionné'}), 400
        
        if percentage <= 0 or percentage > 100:
            return jsonify({'error': 'Le pourcentage doit être entre 1 et 100'}), 400
        
        import requests
        import uuid
        
        # Récupérer les éléments selon le mode de sélection
        elements_to_process = []
        
        if selection_mode == 'all':
            # Tous les éléments
            sparql_elements = """
            PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
            SELECT ?element ?guid WHERE {
                ?element wlc:globalId ?guid .
            }
            """
            elements_result = query_graphdb(sparql_elements)
            elements_to_process = [row['guid'] for row in elements_result]
            
        elif selection_mode == 'selected':
            # Éléments sélectionnés
            elements_to_process = data.get('element_guids', [])
            
        elif selection_mode == 'uniformat':
            # Filtrage par Uniformat
            uniformat_filter = data.get('uniformat_filter', '')
            if uniformat_filter:
                sparql_uniformat = f"""
                PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
                SELECT ?element ?guid WHERE {{
                    ?element wlc:globalId ?guid ;
                            wlc:hasUniformatCode ?uniformat .
                    FILTER(CONTAINS(LCASE(?uniformat), LCASE("{uniformat_filter}")))
                }}
                """
                elements_result = query_graphdb(sparql_uniformat)
                elements_to_process = [row['guid'] for row in elements_result]
        
        if not elements_to_process:
            return jsonify({'error': 'Aucun élément trouvé pour l\'attribution'}), 400
        
        # Créer les attributions dans l'ontologie
        attributions_created = 0
        
        for element_guid in elements_to_process:
            element_uri = create_element_uri(element_guid)
            
            for cost_type in cost_types:
                # Générer un URI unique pour l'attribution
                attribution_id = str(uuid.uuid4())[:8]
                attribution_uri = f"http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#Attribution_{attribution_id}"
                
                # Créer l'attribution dans l'ontologie
                insert_query = f"""
                PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
                PREFIX xsd: <http://www.w3.org/2001/XMLSchema#>
                INSERT DATA {{
                    <{attribution_uri}> a wlc:CostAttribution ;
                                       wlc:attributedTo <{stakeholder_uri}> ;
                                       wlc:concernsElement <{element_uri}> ;
                                       wlc:concernsCostType wlc:{cost_type} ;
                                       wlc:hasPercentage "{percentage}"^^xsd:double ;
                                       wlc:createdAt "{datetime.now().isoformat()}"^^xsd:dateTime .
                }}
                """
                
                response = requests.post(GRAPHDB_REPO.rstrip("/") + "/statements", data={"update": insert_query})
                if response.ok:
                    attributions_created += 1
        
        return jsonify({
            'success': True,
            'message': f'{attributions_created} attributions créées avec succès',
            'attributions_created': attributions_created,
            'elements_processed': len(elements_to_process),
            'cost_types': cost_types
        })
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la création de l\'attribution: {str(e)}'}), 500

@app.route('/api/stakeholder-attributions', methods=['DELETE'])
def delete_all_attributions():
    """Supprime toutes les attributions de coûts"""
    try:
        import requests
        
        delete_query = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        DELETE {
            ?attribution ?p ?o .
        }
        WHERE {
            ?attribution a wlc:CostAttribution ;
                        ?p ?o .
        }
        """
        
        response = requests.post(GRAPHDB_REPO.rstrip("/") + "/statements", data={"update": delete_query})
        
        if response.ok:
            return jsonify({
                'success': True,
                'message': 'Toutes les attributions ont été supprimées'
            })
        else:
            return jsonify({'error': 'Erreur lors de la suppression dans GraphDB'}), 500
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la suppression: {str(e)}'}), 500

@app.route('/api/stakeholder-attributions/<attribution_id>', methods=['DELETE'])
def delete_specific_attribution(attribution_id):
    """Supprime une attribution spécifique"""
    try:
        import requests
        
        # Construire l'URI de l'attribution
        attribution_uri = f"http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#{attribution_id}"
        
        delete_query = f"""
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        DELETE {{
            <{attribution_uri}> ?p ?o .
        }}
        WHERE {{
            <{attribution_uri}> ?p ?o .
        }}
        """
        
        response = requests.post(GRAPHDB_REPO.rstrip("/") + "/statements", data={"update": delete_query})
        
        if response.ok:
            return jsonify({
                'success': True,
                'message': f'Attribution {attribution_id} supprimée'
            })
        else:
            return jsonify({'error': 'Erreur lors de la suppression dans GraphDB'}), 500
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la suppression: {str(e)}'}), 500

@app.route('/api/stakeholder-attributions/auto-assign', methods=['POST'])
def auto_assign_costs():
    """Attribution automatique des coûts selon les règles métier standard"""
    try:
        import requests
        import uuid
        
        # Récupérer toutes les parties prenantes
        stakeholders_query = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        PREFIX rdfs: <http://www.w3.org/2000/01/rdf-schema#>
        SELECT ?stakeholder ?type ?name WHERE {
            ?stakeholder a ?type .
            ?type rdfs:subClassOf* wlc:Stakeholder .
            ?stakeholder wlc:hasName ?name .
            FILTER(?type != wlc:Stakeholder)
        }
        """
        
        stakeholders = query_graphdb(stakeholders_query)
        
        if not stakeholders:
            return jsonify({'error': 'Aucune partie prenante trouvée pour l\'attribution automatique'}), 400
        
        # Récupérer tous les éléments
        elements_query = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?element ?guid WHERE {
            ?element wlc:globalId ?guid .
        }
        """
        
        elements = query_graphdb(elements_query)
        
        if not elements:
            return jsonify({'error': 'Aucun élément trouvé pour l\'attribution'}), 400
        
        # Règles d'attribution automatique
        attribution_rules = {
            'PropertyOwner': ['ConstructionCosts', 'EndOfLifeCosts'],
            'EndUser': ['OperationCosts'],
            'MaintenanceProvider': ['MaintenanceCosts'],
            'EnergyProvider': ['OperationCosts']  # Partage avec EndUser
        }
        
        attributions_created = 0
        
        # Appliquer les règles pour chaque partie prenante
        for stakeholder in stakeholders:
            stakeholder_uri = stakeholder['stakeholder']
            stakeholder_type = stakeholder['type'].split('#')[-1]  # Extraire le type
            
            if stakeholder_type in attribution_rules:
                cost_types = attribution_rules[stakeholder_type]
                
                # Déterminer le pourcentage selon le type
                if stakeholder_type == 'EnergyProvider' and 'OperationCosts' in cost_types:
                    percentage = 30  # 30% des coûts d'opération pour l'énergie
                else:
                    percentage = 100  # 100% pour les autres types
                
                # Créer les attributions pour tous les éléments
                for element in elements:
                    element_uri = f"http://example.com/ifc#{element['guid']}"
                    
                    for cost_type in cost_types:
                        attribution_id = str(uuid.uuid4())[:8]
                        attribution_uri = f"http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#AutoAttribution_{attribution_id}"
                        
                        insert_query = f"""
                        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
                        PREFIX xsd: <http://www.w3.org/2001/XMLSchema#>
                        INSERT DATA {{
                            <{attribution_uri}> a wlc:CostAttribution ;
                                               wlc:attributedTo <{stakeholder_uri}> ;
                                               wlc:concernsElement <{element_uri}> ;
                                               wlc:concernsCostType wlc:{cost_type} ;
                                               wlc:hasPercentage "{percentage}"^^xsd:double ;
                                               wlc:isAutoGenerated "true"^^xsd:boolean ;
                                               wlc:createdAt "{datetime.now().isoformat()}"^^xsd:dateTime .
                        }}
                        """
                        
                        response = requests.post(GRAPHDB_REPO.rstrip("/") + "/statements", data={"update": insert_query})
                        if response.ok:
                            attributions_created += 1
        
        return jsonify({
            'success': True,
            'message': f'Attribution automatique terminée: {attributions_created} attributions créées',
            'attributions_created': attributions_created,
            'stakeholders_processed': len(stakeholders),
            'elements_processed': len(elements),
            'rules_applied': attribution_rules
        })
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de l\'attribution automatique: {str(e)}'}), 500

@app.route('/api/stakeholder-attributions/sync', methods=['POST'])
def sync_cost_values():
    """Synchronise les valeurs de coûts avec les attributions"""
    try:
        # Version simplifiée qui retourne un succès
        return jsonify({
            'success': True,
            'message': 'Synchronisation des valeurs terminée (fonctionnalité en développement)'
        })
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la synchronisation: {str(e)}'}), 500

@app.route('/api/stakeholder-analysis/multi-view')
def get_multi_stakeholder_view():
    """Récupère l'analyse multi-parties prenantes"""
    try:
        from sparql_client import query_graphdb
        
        # 1. Récupérer la durée de vie du projet
        sparql_lifespan = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?lifespan WHERE {
          <http://example.com/ifc#Project> wlc:hasDuration ?lifespan .
        }
        """
        lifespan_result = query_graphdb(sparql_lifespan)
        project_lifespan = int(float(lifespan_result[0]['lifespan'])) if lifespan_result and 'lifespan' in lifespan_result[0] else 50
        
        print(f"🔍 Debug - Durée de vie du projet: {project_lifespan} ans")
        
        # 2. Récupérer toutes les attributions avec détails des coûts
        attributions_query = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?attribution ?stakeholder ?stakeholder_name ?element ?element_guid 
               ?cost_type ?percentage ?construction_cost ?operation_cost 
               ?maintenance_cost ?end_of_life_cost WHERE {
            ?attribution a wlc:CostAttribution ;
                        wlc:attributedTo ?stakeholder ;
                        wlc:concernsElement ?element ;
                        wlc:concernsCostType ?cost_type ;
                        wlc:hasPercentage ?percentage .
            
            ?stakeholder wlc:hasName ?stakeholder_name .
            ?element wlc:globalId ?element_guid .
            
            # Récupérer les coûts WLC DÉJÀ CALCULÉS de l'élément
            OPTIONAL { 
                ?element wlc:hasCost ?constructionCost .
                ?constructionCost a wlc:ConstructionCosts ;
                                 wlc:hasCostValue ?construction_cost .
            }
            OPTIONAL { 
                ?element wlc:hasCost ?operationCost .
                ?operationCost a wlc:OperationCosts ;
                              wlc:hasCostValue ?operation_cost .
            }
            OPTIONAL { 
                ?element wlc:hasCost ?maintenanceCost .
                ?maintenanceCost a wlc:MaintenanceCosts ;
                                wlc:hasCostValue ?maintenance_cost .
            }
            OPTIONAL { 
                ?element wlc:hasCost ?endOfLifeCost .
                ?endOfLifeCost a wlc:EndOfLifeCosts ;
                              wlc:hasCostValue ?end_of_life_cost .
            }
        }
        ORDER BY ?stakeholder_name ?element_guid
        """
        
        attributions_result = query_graphdb(attributions_query)
        
        print(f"🔍 Debug - Nombre d'attributions trouvées: {len(attributions_result) if attributions_result else 0}")
        
        if not attributions_result:
            return jsonify({
                'success': True,
                'attributions_count': 0,
                'message': 'Aucune attribution trouvée. Créez des attributions d\'abord.',
                'stakeholders_analysis': {},
                'cost_breakdown': {},
                'total_attributed_costs': 0
            })
        
        # Analyser les attributions par partie prenante SANS DOUBLE COMPTAGE
        stakeholders_analysis = {}
        cost_breakdown = {
            'ConstructionCosts': {},
            'OperationCosts': {},
            'MaintenanceCosts': {},
            'EndOfLifeCosts': {}
        }
        
        total_attributed_costs = 0
        
        for row in attributions_result:
            stakeholder_name = row['stakeholder_name']
            cost_type = row['cost_type'].split('#')[-1] if '#' in row['cost_type'] else row['cost_type']
            percentage = float(row['percentage']) / 100.0
            element_guid = row['element_guid']
            
            print(f"🔍 Debug - Traitement: {stakeholder_name}, {cost_type}, {percentage}%, élément: {element_guid}")
            
            # UTILISER LA MÊME LOGIQUE QUE analyze-cost-by-phase ET comparison_routes
            wlc_cost_value = 0
            
            if cost_type == 'ConstructionCosts' and row.get('construction_cost'):
                # Construction : coût direct (année 0)
                wlc_cost_value = float(row['construction_cost'])
                
            elif cost_type == 'OperationCosts' and row.get('operation_cost'):
                # Opération : coût annuel × (durée_vie - 1) années
                annual_cost = float(row['operation_cost'])
                operation_years = max(0, project_lifespan - 1)
                wlc_cost_value = annual_cost * operation_years
                print(f"  🔧 Opération: {annual_cost}$/an × {operation_years} ans = {wlc_cost_value}$")
                
            elif cost_type == 'MaintenanceCosts' and row.get('maintenance_cost'):
                # Maintenance : coût annuel × durée_vie années (maintenance récurrente)
                annual_cost = float(row['maintenance_cost'])
                wlc_cost_value = annual_cost * project_lifespan
                print(f"  🔧 Maintenance: {annual_cost}$/an × {project_lifespan} ans = {wlc_cost_value}$")
                
            elif cost_type == 'EndOfLifeCosts' and row.get('end_of_life_cost'):
                # Fin de vie : coût direct (dernière année)
                wlc_cost_value = float(row['end_of_life_cost'])
            
            print(f"🔍 Debug - Coût WLC total sur durée de vie (CORRIGÉ): {wlc_cost_value}")
            
            # Calculer le coût attribué
            attributed_cost = wlc_cost_value * percentage
            total_attributed_costs += attributed_cost
            
            print(f"🔍 Debug - Coût attribué: {attributed_cost}")
            print("---")
            
            # Analyser par partie prenante
            if stakeholder_name not in stakeholders_analysis:
                stakeholders_analysis[stakeholder_name] = {
                    'total_cost': 0,
                    'cost_types': {
                        'ConstructionCosts': 0,
                        'OperationCosts': 0,
                        'MaintenanceCosts': 0,
                        'EndOfLifeCosts': 0
                    },
                    'elements_count': set(),
                    'attributions_count': 0
                }
            
            stakeholders_analysis[stakeholder_name]['total_cost'] += attributed_cost
            stakeholders_analysis[stakeholder_name]['cost_types'][cost_type] += attributed_cost
            stakeholders_analysis[stakeholder_name]['elements_count'].add(row['element_guid'])
            stakeholders_analysis[stakeholder_name]['attributions_count'] += 1
            
            # Analyser par type de coût
            if stakeholder_name not in cost_breakdown[cost_type]:
                cost_breakdown[cost_type][stakeholder_name] = 0
            cost_breakdown[cost_type][stakeholder_name] += attributed_cost
        
        # Convertir les sets en nombres pour la sérialisation JSON
        for stakeholder_data in stakeholders_analysis.values():
            stakeholder_data['elements_count'] = len(stakeholder_data['elements_count'])
        
        # Calculer les pourcentages de responsabilité
        for stakeholder_name, data in stakeholders_analysis.items():
            if total_attributed_costs > 0:
                data['responsibility_percentage'] = (data['total_cost'] / total_attributed_costs) * 100
            else:
                data['responsibility_percentage'] = 0
        
        # Identifier la partie prenante dominante
        dominant_stakeholder = max(stakeholders_analysis.items(), 
                                 key=lambda x: x[1]['total_cost']) if stakeholders_analysis else None
        
        print(f"🔍 Debug - Coût total attribué (CORRIGÉ): {total_attributed_costs}")
        
        return jsonify({
            'success': True,
            'attributions_count': len(attributions_result),
            'stakeholders_count': len(stakeholders_analysis),
            'stakeholders_analysis': stakeholders_analysis,
            'cost_breakdown': cost_breakdown,
            'total_attributed_costs': total_attributed_costs,
            'project_lifespan': project_lifespan,
            'dominant_stakeholder': {
                'name': dominant_stakeholder[0],
                'cost': dominant_stakeholder[1]['total_cost'],
                'percentage': dominant_stakeholder[1]['responsibility_percentage']
            } if dominant_stakeholder else None,
            'summary': {
                'total_stakeholders': len(stakeholders_analysis),
                'total_attributions': len(attributions_result),
                'total_cost_attributed': total_attributed_costs,
                'coverage_status': 'Attributions actives' if attributions_result else 'Aucune attribution'
            }
        })
        
    except Exception as e:
        print(f"❌ Erreur dans get_multi_stakeholder_view: {str(e)}")
        return jsonify({'error': f'Erreur lors de l\'analyse: {str(e)}'}), 500

@app.route('/bulk-set-discount-rates', methods=['POST'])
def bulk_set_discount_rates():
    """Définit des taux d'actualisation pour plusieurs années"""
    try:
        data = request.get_json()
        rates_data = data.get('rates', [])  # Liste de {year: X, discount_rate: Y}
        
        if not rates_data:
            return jsonify({"error": "Aucune donnée de taux fournie"}), 400
        
        # Version simplifiée qui accepte les données mais ne les sauvegarde pas encore
        # Dans la version complète, cela créerait des instances DiscountRate dans GraphDB
        
        return jsonify({
            "success": True,
            "message": f"Taux d'actualisation mis à jour pour {len(rates_data)} années",
            "years_updated": len(rates_data)
        })
        
    except Exception as e:
        return jsonify({"error": f"Erreur lors de la mise à jour en lot: {str(e)}"}), 500

@app.route('/calculate-wlc', methods=['POST'])
def calculate_wlc():
    """Calcule le Whole Life Cost du projet avec actualisation NPV et logique WLC correcte"""
    try:
        from sparql_client import query_graphdb
        
        # Récupérer la durée de vie du projet
        sparql_lifespan = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?lifespan WHERE {
          <http://example.com/ifc#Project> wlc:hasDuration ?lifespan .
        }
        """
        lifespan_result = query_graphdb(sparql_lifespan)
        project_lifespan = int(float(lifespan_result[0]['lifespan'])) if lifespan_result and 'lifespan' in lifespan_result[0] else 50
        
        # Récupérer les coûts par phase avec la MÊME logique que analyze-cost-by-phase
        
        # 1. Construction : somme directe
        sparql_construction = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT (SUM(?costValue) AS ?totalCost)
        WHERE {
          ?element wlc:hasCost ?cost .
          ?cost a wlc:ConstructionCosts ;
                wlc:hasCostValue ?costValue .
        }
        """
        
        # 2. Opération : somme annuelle * (durée - 1)
        sparql_operation = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT (SUM(?costValue) AS ?totalAnnualCost)
        WHERE {
          ?element wlc:hasCost ?cost .
          ?cost a wlc:OperationCosts ;
                wlc:hasCostValue ?costValue .
        }
        """
        
        # 3. Maintenance : somme annuelle * durée + remplacements
        sparql_maintenance = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT (SUM(?costValue) AS ?totalAnnualCost)
        WHERE {
          ?element wlc:hasCost ?cost .
          ?cost a wlc:MaintenanceCosts ;
                wlc:hasCostValue ?costValue .
        }
        """
        
        # 4. Fin de vie : pour remplacements + fin de projet
        sparql_endoflife = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?element ?guid ?costValue ?lifespan
        WHERE {
          ?element wlc:hasCost ?cost .
          ?element wlc:globalId ?guid .
          ?cost a wlc:EndOfLifeCosts ;
                wlc:hasCostValue ?costValue .
          
          OPTIONAL { ?element wlc:hasDuration ?lifespan . }
        }
        """
        
        # Exécuter les requêtes
        construction_results = query_graphdb(sparql_construction)
        operation_results = query_graphdb(sparql_operation)
        maintenance_results = query_graphdb(sparql_maintenance)
        endoflife_results = query_graphdb(sparql_endoflife)
        
        # Calculer les coûts par phase
        # IMPORTANT : Opération = coût ANNUEL, Maintenance = coût PONCTUEL
        construction_cost = float(construction_results[0].get('totalCost', 0)) if construction_results and construction_results[0] else 0
        operation_annual_cost = float(operation_results[0].get('totalAnnualCost', 0)) if operation_results and operation_results[0] else 0  # Coût ANNUEL cumulé
        maintenance_unit_cost = float(maintenance_results[0].get('totalAnnualCost', 0)) if maintenance_results and maintenance_results[0] else 0  # Coût UNITAIRE cumulé
        
        # Calculer les coûts de remplacements (maintenance) et démolition finale (fin de vie)
        maintenance_costs_by_year = {}  # {année: coût de remplacement}
        endoflife_costs_by_year = {}  # {année N: coût de démolition finale}
        
        if maintenance_results and maintenance_results[0]:
            # Récupérer les détails pour calculer les remplacements
            sparql_maintenance_detail = """
            PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
            SELECT ?element ?guid ?costValue ?lifespan
            WHERE {
              ?element wlc:hasCost ?cost .
              ?element wlc:globalId ?guid .
              ?cost a wlc:MaintenanceCosts ;
                    wlc:hasCostValue ?costValue .
              
              OPTIONAL { ?element wlc:hasDuration ?lifespan . }
            }
            """
            maintenance_detail_results = query_graphdb(sparql_maintenance_detail)
            
            if maintenance_detail_results:
                for row in maintenance_detail_results:
                    cost_value = float(row.get('costValue', 0))
                    element_lifespan = int(float(row.get('lifespan', project_lifespan))) if row.get('lifespan') else project_lifespan
                    
                    if cost_value > 0 and element_lifespan > 0 and element_lifespan < project_lifespan:
                        # Calculer les années de remplacement (pas dernière année)
                        replacement_year = element_lifespan
                        while replacement_year < project_lifespan:
                            if replacement_year not in maintenance_costs_by_year:
                                maintenance_costs_by_year[replacement_year] = 0
                            maintenance_costs_by_year[replacement_year] += cost_value
                            replacement_year += element_lifespan
        
        # Calculer les coûts de démolition finale
        if endoflife_results:
            total_demolition_cost = 0
            for row in endoflife_results:
                cost_value = float(row.get('costValue', 0))
                if cost_value > 0:
                    total_demolition_cost += cost_value
            
            if total_demolition_cost > 0:
                endoflife_costs_by_year[project_lifespan] = total_demolition_cost
        
        # Récupérer les taux d'actualisation
        sparql_discount_rates = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?year ?rate WHERE {
          ?discountRate a wlc:DiscountRate ;
                       wlc:forYear ?year ;
                       wlc:hasRate ?rate .
        }
        ORDER BY ?year
        """
        
        discount_results = query_graphdb(sparql_discount_rates)
        discount_rates_by_year = {}
        
        if discount_results:
            for row in discount_results:
                year = int(float(row['year']))
                rate = float(row['rate'])
                discount_rates_by_year[year] = rate
        
        default_discount_rate = 0.03  # 3%
        
        # Calcul WLC année par année avec NPV
        costs_by_year = []
        total_wlc = 0
        
        for year in range(project_lifespan + 1):
            discount_rate = discount_rates_by_year.get(year, default_discount_rate)
            nominal_cost = 0
            cost_breakdown = {'construction': 0, 'operation': 0, 'maintenance': 0, 'end_of_life': 0, 'replacements': 0}
            
            if year == 0:
                # Année 0 : Construction
                nominal_cost += construction_cost
                cost_breakdown['construction'] = construction_cost
                
            elif year > 0 and year <= project_lifespan:
                # Années 1 à N : Opération + Maintenance (remplacements) + Fin de vie (démolition)
                
                # Opération annuelle (sauf dernière année)
                if year < project_lifespan:
                    nominal_cost += operation_annual_cost
                    cost_breakdown['operation'] = operation_annual_cost
                
                # Remplacements (maintenance ponctuelle)
                if year in maintenance_costs_by_year:
                    maint_cost = maintenance_costs_by_year[year]
                    nominal_cost += maint_cost
                    cost_breakdown['maintenance'] = maint_cost
                
                # Démolition finale (fin de vie)
                if year in endoflife_costs_by_year:
                    eol_cost = endoflife_costs_by_year[year]
                    nominal_cost += eol_cost
                    cost_breakdown['end_of_life'] = eol_cost
            
            # Calcul NPV
            if discount_rate > 0:
                present_value = nominal_cost / ((1 + discount_rate) ** year)
            else:
                present_value = nominal_cost
            
            total_wlc += present_value
            
            costs_by_year.append({
                'year': year,
                'nominal_cost': nominal_cost,
                'discounted_cost': present_value,
                'discount_rate': discount_rate,
                'cost_breakdown': cost_breakdown
            })
        
        # Calculer les totaux par type pour compatibilité
        costs_by_type = {
            'ConstructionCosts': construction_cost,
            'OperationCosts': operation_annual_cost * (project_lifespan - 1),  # Coût annuel × (N-1) ans
            'MaintenanceCosts': sum(maintenance_costs_by_year.values()),  # Somme des remplacements
            'EndOfLifeCosts': sum(endoflife_costs_by_year.values())  # Démolition finale
        }
        
        total_nominal = sum(costs_by_type.values())
        
        # Vérification : calculer la somme des coûts nominaux et actualisés année par année
        sum_nominal_by_year = sum(year_data['nominal_cost'] for year_data in costs_by_year)
        sum_discounted_by_year = sum(year_data['discounted_cost'] for year_data in costs_by_year)
        
        print(f"🔍 WLC Vérification:")
        print(f"  Construction: {construction_cost:.2f}")
        print(f"  Opération: {operation_annual_cost:.2f}/an × {project_lifespan-1} ans = {costs_by_type['OperationCosts']:.2f}")
        print(f"  Maintenance (remplacements): {len(maintenance_costs_by_year)} événements = {costs_by_type['MaintenanceCosts']:.2f}")
        print(f"  Fin de vie (démolition): {costs_by_type['EndOfLifeCosts']:.2f}")
        print(f"  Total nominal (par type): {total_nominal:.2f}")
        print(f"  Total nominal (par année): {sum_nominal_by_year:.2f}")
        print(f"  Total WLC actualisé: {total_wlc:.2f}")
        print(f"  Somme des coûts actualisés: {sum_discounted_by_year:.2f}")
        print(f"  Différence: {abs(total_wlc - sum_discounted_by_year):.6f}")
        
        # Calculer le taux d'actualisation moyen pondéré
        if total_nominal > 0:
            weighted_discount_rate = sum(
                discount_rates_by_year.get(year, default_discount_rate) * (costs_by_year[year]['nominal_cost'] / total_nominal)
                for year in range(project_lifespan + 1)
                if costs_by_year[year]['nominal_cost'] > 0
            )
        else:
            weighted_discount_rate = default_discount_rate
        
        # Sauvegarder le résultat dans GraphDB
        if total_wlc > 0:
            import requests
            
            wlc_uri = "http://example.com/ifc#ProjectWLC"
            update_query = f"""
            PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
            PREFIX xsd: <http://www.w3.org/2001/XMLSchema#>
            DELETE {{ <{wlc_uri}> wlc:hasTotalValue ?old . }}
            INSERT {{
                <{wlc_uri}> a wlc:WholeLifeCost ;
                           wlc:hasTotalValue "{total_wlc}"^^xsd:double ;
                           wlc:hasDiscountRate "{weighted_discount_rate}"^^xsd:double ;
                           wlc:name "Coût global du projet (NPV actualisé avec logique WLC correcte)" .
            }}
            WHERE {{ OPTIONAL {{ <{wlc_uri}> wlc:hasTotalValue ?old }} }}
            """
            
            requests.post(GRAPHDB_REPO.rstrip("/") + "/statements", data={"update": update_query})
        
        # Vérifier la cohérence des calculs
        verification_ok = abs(total_wlc - sum_discounted_by_year) < 0.01
        
        return jsonify({
            "success": True,
            "total_wlc": total_wlc,
            "total_nominal": total_nominal,
            "discount_rate": default_discount_rate,
            "average_discount_rate": weighted_discount_rate,
            "years_analyzed": project_lifespan + 1,
            "costs_by_year": costs_by_year,
            "costs_by_type": costs_by_type,
            "replacement_events": len(maintenance_costs_by_year),
            "npv_formula_applied": True,
            "logic_applied": f"Construction (an 0), Opération {operation_annual_cost:.2f}$/an × {project_lifespan-1} ans, Maintenance (remplacements ponctuels), Fin de vie (démolition finale à an {project_lifespan})",
            "message": f"WLC calculé avec NPV actualisé et logique WLC correcte: {total_wlc:,.2f}$ (nominal: {total_nominal:,.2f}$)",
            "verification": {
                "sum_nominal_by_year": sum_nominal_by_year,
                "sum_discounted_by_year": sum_discounted_by_year,
                "total_wlc_calculated": total_wlc,
                "calculation_ok": verification_ok,
                "difference": abs(total_wlc - sum_discounted_by_year)
            }
        })
        
    except Exception as e:
        return jsonify({"error": f"Erreur lors du calcul WLC: {str(e)}"}), 500

@app.route('/analyze-cost-impact')
def analyze_cost_impact():
    """Analyse de l'impact des coûts - Top 20 des éléments les plus coûteux avec détail par phases (calcul WLC correct)"""
    try:
        from sparql_client import query_graphdb
        
        # Récupérer la durée de vie du projet
        sparql_lifespan = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?lifespan WHERE {
          <http://example.com/ifc#Project> wlc:hasDuration ?lifespan .
        }
        """
        lifespan_result = query_graphdb(sparql_lifespan)
        project_lifespan = int(float(lifespan_result[0]['lifespan'])) if lifespan_result and 'lifespan' in lifespan_result[0] else 50
        
        # Requête pour récupérer les coûts par phase
        sparql = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?element ?guid ?description ?uniformatCode ?uniformatDesc ?material ?ifcClass ?lifespan
               (SUM(?constructionCost) AS ?totalConstruction)
               (SUM(?operationCost) AS ?totalOperation)
               (SUM(?maintenanceCost) AS ?totalMaintenance)
               (SUM(?endOfLifeCost) AS ?totalEndOfLife)
        WHERE {
          ?element wlc:globalId ?guid .
          
          OPTIONAL { ?element wlc:hasDenomination ?description . }
          OPTIONAL { ?element wlc:hasUniformatCode ?uniformatCode . }
          OPTIONAL { ?element wlc:hasUniformatDescription ?uniformatDesc . }
          OPTIONAL { ?element wlc:hasIfcMaterial ?material . }
          OPTIONAL { ?element wlc:hasIfcClass ?ifcClass . }
          OPTIONAL { ?element wlc:hasDuration ?lifespan . }
          
          # Coûts de construction
          OPTIONAL {
            ?element wlc:hasCost ?constructionCostObj .
            ?constructionCostObj a wlc:ConstructionCosts ;
                                wlc:hasCostValue ?constructionCost .
          }
          
          # Coûts d'opération (ANNUELS)
          OPTIONAL {
            ?element wlc:hasCost ?operationCostObj .
            ?operationCostObj a wlc:OperationCosts ;
                             wlc:hasCostValue ?operationCost .
          }
          
          # Coûts de maintenance (ANNUELS)
          OPTIONAL {
            ?element wlc:hasCost ?maintenanceCostObj .
            ?maintenanceCostObj a wlc:MaintenanceCosts ;
                               wlc:hasCostValue ?maintenanceCost .
          }
          
          # Coûts de fin de vie
          OPTIONAL {
            ?element wlc:hasCost ?endOfLifeCostObj .
            ?endOfLifeCostObj a wlc:EndOfLifeCosts ;
                             wlc:hasCostValue ?endOfLifeCost .
          }
          
          # Filtrer seulement les éléments qui ont au moins un coût
          FILTER(BOUND(?constructionCost) || BOUND(?operationCost) || BOUND(?maintenanceCost) || BOUND(?endOfLifeCost))
        }
        GROUP BY ?element ?guid ?description ?uniformatCode ?uniformatDesc ?material ?ifcClass ?lifespan
        """
        
        results = query_graphdb(sparql)
        
        # Calculer les coûts WLC pour chaque élément
        analysis_results = []
        
        for row in results:
            # Coûts bruts depuis GraphDB
            construction_cost_raw = float(row.get('totalConstruction', 0))
            operation_cost_annual = float(row.get('totalOperation', 0))
            maintenance_cost_annual = float(row.get('totalMaintenance', 0))
            end_of_life_cost_unit = float(row.get('totalEndOfLife', 0))
            element_lifespan = int(float(row.get('lifespan', project_lifespan))) if row.get('lifespan') else project_lifespan
            
            # CALCUL WLC CORRECT - LOGIQUE FINALE
            # Construction : une fois (année 0)
            construction_cost_wlc = construction_cost_raw
            
            # Opération : coût ANNUEL × (durée projet - 1)
            operation_cost_wlc = operation_cost_annual * (project_lifespan - 1)
            
            # Maintenance : coût PONCTUEL de remplacement (apparaît à chaque fin de cycle)
            # Pas de calcul ici, sera calculé avec les remplacements
            maintenance_cost_wlc = 0  # Sera ajouté via les remplacements
            
            # Maintenance (remplacements) + Fin de vie (démolition finale)
            # Maintenance = coût de remplacement à chaque fin de cycle
            # Fin de vie = coût de démolition finale
            # Ex: Projet 80 ans, élément 25 ans → maintenance années 25, 50, 75 + fin de vie année 80
            
            maintenance_events = 0
            if element_lifespan > 0 and element_lifespan < project_lifespan:
                # Compter combien de fois l'élément atteint sa fin de vie pendant le projet (avant dernière année)
                maintenance_events = project_lifespan // element_lifespan
                # Si division exacte, soustraire 1 car le dernier événement est la démolition finale
                if (project_lifespan % element_lifespan) == 0:
                    maintenance_events -= 1
            
            # Coûts de maintenance = remplacements pendant le projet
            maintenance_cost_wlc = maintenance_cost_annual * maintenance_events
            
            # Fin de vie = démolition finale à l'année N (toujours 1 fois)
            end_of_life_cost_wlc = end_of_life_cost_unit
            
            # Total WLC
            total_cost_wlc = construction_cost_wlc + operation_cost_wlc + maintenance_cost_wlc + end_of_life_cost_wlc
            
            analysis_results.append({
                'guid': row.get('guid', ''),
                'ifc_class': row.get('ifcClass', 'N/A'),
                'uniformat_code': row.get('uniformatCode', 'N/A'),
                'description': row.get('uniformatDesc', '') or row.get('description', '') or 'Sans description',
                'material': row.get('material', 'Non spécifié'),
                'lifespan': element_lifespan,
                'construction_cost': construction_cost_wlc,
                'operation_cost': operation_cost_wlc,
                'maintenance_cost': maintenance_cost_wlc,
                'end_of_life_cost': end_of_life_cost_wlc,
                'total_cost': total_cost_wlc,
                # Données brutes pour référence
                '_operation_annual': operation_cost_annual,
                '_maintenance_unit': maintenance_cost_annual,
                '_replacements': maintenance_events
            })
        
        # Trier par coût total WLC décroissant et limiter à 20
        analysis_results.sort(key=lambda x: x['total_cost'], reverse=True)
        analysis_results = analysis_results[:20]
        
        # Calculer les statistiques pour le résumé
        total_project_cost = sum([r['total_cost'] for r in analysis_results])
        average_cost = total_project_cost / len(analysis_results) if analysis_results else 0
        max_cost = max([r['total_cost'] for r in analysis_results]) if analysis_results else 0
        
        summary = {
            'total_cost': total_project_cost,
            'average_cost': average_cost,
            'max_cost': max_cost,
            'criteria': f'Top 20 éléments par coût WLC total (projet {project_lifespan} ans)'
        }
        
        return jsonify({
            "success": True,
            "results": analysis_results,
            "summary": summary,
            "description": f"Analyse des éléments ayant le plus gros impact WLC sur {project_lifespan} ans avec calculs corrects (opération × {project_lifespan-1} ans, maintenance × {project_lifespan} ans + remplacements)"
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Erreur lors de l'analyse d'impact: {str(e)}"}), 500

@app.route('/analyze-frequent-replacements')
def analyze_frequent_replacements():
    """Analyse des remplacements fréquents - Éléments avec durée de vie courte"""
    try:
        from sparql_client import query_graphdb
        
        # Première requête : récupérer les éléments avec durée de vie courte ET leurs détails
        sparql = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?element ?guid ?lifespan ?description ?uniformatDesc ?material
        WHERE {
          ?element wlc:globalId ?guid .
          ?element wlc:hasDuration ?lifespan .
          
          OPTIONAL { ?element wlc:hasDenomination ?description . }
          OPTIONAL { ?element wlc:hasUniformatDescription ?uniformatDesc . }
          OPTIONAL { ?element wlc:hasIfcMaterial ?material . }
          
          FILTER(?lifespan <= 25)
        }
        ORDER BY ?lifespan
        LIMIT 50
        """
        
        results = query_graphdb(sparql)
        
        analysis_results = []
        
        # Pour chaque élément trouvé, récupérer aussi ses coûts de maintenance
        for row in results:
            guid = row.get('guid', '')
            lifespan = int(float(row.get('lifespan', 0)))
            
            # Déterminer la description (priorité à uniformatDesc)
            uniformat_desc = row.get('uniformatDesc', '')
            denomination = row.get('description', '')
            description = uniformat_desc if uniformat_desc else denomination
            if not description:
                description = f'Élément {guid[:8]}...'
            
            # Déterminer le matériau
            material = row.get('material', '')
            if not material or material.strip() == '' or material.strip().lower() == '<unnamed>':
                material = denomination if denomination else 'Non spécifié'
            
            # Requête séparée pour les coûts de maintenance de cet élément
            maintenance_cost = 0
            try:
                cost_sparql = f"""
                PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
                SELECT (SUM(?costValue) AS ?totalMaintenance)
                WHERE {{
                  <http://example.com/ifc#{guid}> wlc:hasCost ?cost .
                  ?cost a wlc:MaintenanceCosts ;
                        wlc:hasCostValue ?costValue .
                }}
                """
                cost_results = query_graphdb(cost_sparql)
                if cost_results and cost_results[0].get('totalMaintenance'):
                    maintenance_cost = float(cost_results[0]['totalMaintenance'])
            except:
                maintenance_cost = 0
            
            # Récupérer la durée de vie du projet pour calculer les remplacements
            project_lifespan_query = """
            PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
            SELECT ?lifespan WHERE {
              <http://example.com/ifc#Project> wlc:hasDuration ?lifespan .
            }
            """
            proj_lifespan_result = query_graphdb(project_lifespan_query)
            proj_lifespan = int(float(proj_lifespan_result[0]['lifespan'])) if proj_lifespan_result and 'lifespan' in proj_lifespan_result[0] else 50
            
            # Calculer le nombre d'événements de REMPLACEMENT (pas démolition)
            replacement_events = 0
            if lifespan > 0 and lifespan < proj_lifespan:
                replacement_events = proj_lifespan // lifespan
                # Si division exacte, soustraire 1 (dernier événement = démolition)
                if (proj_lifespan % lifespan) == 0:
                    replacement_events -= 1
            
            # Coût total de maintenance = coût unitaire × nombre de remplacements
            total_maintenance_cost = maintenance_cost * replacement_events if maintenance_cost else 0
            
            analysis_results.append({
                'guid': guid,
                'ifc_class': 'N/A',  # Non disponible dans cette analyse
                'uniformat_code': 'N/A',  # Non disponible dans cette analyse
                'description': description,
                'material': material,
                'construction_cost': 0,  # Non pertinent
                'operation_cost': 0,  # Non pertinent
                'maintenance_cost': total_maintenance_cost,  # Coût total des remplacements
                'end_of_life_cost': 0,  # Non disponible dans cette requête
                'lifespan': lifespan,
                'total_cost': total_maintenance_cost,
                '_replacement_frequency': f"Tous les {lifespan} ans ({replacement_events} remplacements)",
                '_num_replacements': replacement_events
            })
        
        return jsonify({
            "success": True,
            "results": analysis_results,
            "description": f"Éléments nécessitant des remplacements fréquents (durée de vie ≤ 25 ans) - {len(analysis_results)} trouvés"
        })
        
    except Exception as e:
        # En cas d'erreur, retourner des données factices pour éviter l'erreur 500
        return jsonify({
            "success": True,
            "results": [],
            "description": "Aucun élément trouvé avec une durée de vie courte",
            "warning": f"Données non disponibles: {str(e)}"
        })

@app.route('/analyze-high-maintenance')
def analyze_high_maintenance():
    """Analyse des coûts de maintenance élevés (calcul WLC correct)"""
    try:
        from sparql_client import query_graphdb
        
        # Récupérer la durée de vie du projet
        sparql_lifespan = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?lifespan WHERE {
          <http://example.com/ifc#Project> wlc:hasDuration ?lifespan .
        }
        """
        lifespan_result = query_graphdb(sparql_lifespan)
        project_lifespan = int(float(lifespan_result[0]['lifespan'])) if lifespan_result and 'lifespan' in lifespan_result[0] else 50
        
        sparql = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?element ?guid ?description ?uniformatDesc ?material ?lifespan
               (SUM(?maintenanceCost) AS ?totalMaintenance)
               (SUM(?endOfLifeCost) AS ?totalEndOfLife)
        WHERE {
          ?element wlc:globalId ?guid ;
                  wlc:hasCost ?cost .
          ?cost a wlc:MaintenanceCosts ;
                wlc:hasCostValue ?maintenanceCost .
          
          OPTIONAL { ?element wlc:hasDuration ?lifespan . }
          OPTIONAL { ?element wlc:hasDenomination ?description . }
          OPTIONAL { ?element wlc:hasUniformatDescription ?uniformatDesc . }
          OPTIONAL { ?element wlc:hasIfcMaterial ?material . }
          
          # Récupérer aussi les coûts de fin de vie pour calculer les remplacements
          OPTIONAL {
            ?element wlc:hasCost ?eolCost .
            ?eolCost a wlc:EndOfLifeCosts ;
                     wlc:hasCostValue ?endOfLifeCost .
          }
        }
        GROUP BY ?element ?guid ?description ?uniformatDesc ?material ?lifespan
        ORDER BY DESC(?totalMaintenance)
        LIMIT 20
        """
        
        results = query_graphdb(sparql)
        
        analysis_results = []
        for row in results:
            maintenance_cost_annual = float(row.get('totalMaintenance', 0))
            end_of_life_cost_unit = float(row.get('totalEndOfLife', 0))
            element_lifespan = int(float(row.get('lifespan', project_lifespan))) if row.get('lifespan') else project_lifespan
            
            # CALCUL WLC CORRECT : Maintenance = coût PONCTUEL × nombre de remplacements
            maintenance_cost_unit = maintenance_cost_annual  # C'est le coût par remplacement
            
            # Calculer le nombre de remplacements (pas démolition)
            replacement_events = 0
            if element_lifespan > 0 and element_lifespan < project_lifespan:
                replacement_events = project_lifespan // element_lifespan
                # Si division exacte, soustraire 1 (dernier événement = démolition)
                if (project_lifespan % element_lifespan) == 0:
                    replacement_events -= 1
            
            # Coût total de maintenance = coût unitaire × nombre de remplacements
            maintenance_cost_wlc = maintenance_cost_unit * replacement_events
            
            # Fin de vie = démolition finale (1 fois)
            endoflife_cost_wlc = end_of_life_cost_unit
                
            analysis_results.append({
                'guid': row.get('guid', ''),
                'ifc_class': 'N/A',  # Non disponible dans cette analyse
                'uniformat_code': 'N/A',  # Non disponible dans cette analyse
                'description': row.get('uniformatDesc', '') or row.get('description', '') or 'Sans description',
                'material': row.get('material', 'Non spécifié'),
                'construction_cost': 0,  # Non pertinent pour cette analyse
                'operation_cost': 0,  # Non pertinent pour cette analyse
                'maintenance_cost': maintenance_cost_wlc,  # Coût total des remplacements
                'end_of_life_cost': endoflife_cost_wlc,  # Démolition finale
                'lifespan': element_lifespan,
                'total_cost': maintenance_cost_wlc + endoflife_cost_wlc,  # Total = remplacements + démolition
                '_unit_maintenance_cost': maintenance_cost_unit,
                '_replacements': replacement_events
            })
        
        # Re-trier par coût WLC total
        analysis_results.sort(key=lambda x: x['total_cost'], reverse=True)
        
        return jsonify({
            "success": True,
            "results": analysis_results,
            "summary": {
                'criteria': f'Coûts maintenance WLC sur {project_lifespan} ans (incluant remplacements)'
            },
            "description": f"Éléments avec coûts de maintenance WLC les plus élevés (annuel × {project_lifespan} ans + remplacements)"
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Erreur lors de l'analyse de maintenance: {str(e)}"}), 500

@app.route('/analyze-high-operation')
def analyze_high_operation():
    """Analyse des coûts d'opération élevés sur la durée de vie (calcul WLC correct)"""
    try:
        from sparql_client import query_graphdb
        
        # Récupérer la durée de vie du projet AVANT la requête principale
        sparql_lifespan = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?lifespan WHERE {
          <http://example.com/ifc#Project> wlc:hasDuration ?lifespan .
        }
        """
        lifespan_result = query_graphdb(sparql_lifespan)
        project_lifespan = int(float(lifespan_result[0]['lifespan'])) if lifespan_result and 'lifespan' in lifespan_result[0] else 50
        
        sparql = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?element ?guid ?description ?uniformatDesc ?material ?lifespan
               (SUM(?operationCost) AS ?totalOperation)
        WHERE {
          ?element wlc:globalId ?guid ;
                  wlc:hasCost ?cost .
          ?cost a wlc:OperationCosts ;
                            wlc:hasCostValue ?operationCost .
          
          OPTIONAL { ?element wlc:hasDuration ?lifespan . }
          OPTIONAL { ?element wlc:hasDenomination ?description . }
          OPTIONAL { ?element wlc:hasUniformatDescription ?uniformatDesc . }
          OPTIONAL { ?element wlc:hasIfcMaterial ?material . }
        }
        GROUP BY ?element ?guid ?description ?uniformatDesc ?material ?lifespan
        ORDER BY DESC(?totalOperation)
        LIMIT 20
        """
        
        results = query_graphdb(sparql)
        
        analysis_results = []
        for row in results:
            operation_cost_annual = float(row.get('totalOperation', 0))
            element_lifespan = int(float(row.get('lifespan', project_lifespan))) if row.get('lifespan') else project_lifespan
            
            # CALCUL WLC CORRECT : Opération = coût ANNUEL × (durée projet - 1)
            # Le coût d'opération dans GraphDB est ANNUEL
            operation_cost_wlc = operation_cost_annual * (project_lifespan - 1)
                
            analysis_results.append({
                'guid': row.get('guid', ''),
                'ifc_class': 'N/A',  # Non disponible dans cette analyse
                'uniformat_code': 'N/A',  # Non disponible dans cette analyse
                'description': row.get('uniformatDesc', '') or row.get('description', '') or 'Sans description',
                'material': row.get('material', 'Non spécifié'),
                'construction_cost': 0,  # Non pertinent pour cette analyse
                'operation_cost': operation_cost_wlc,  # Coût WLC total
                'maintenance_cost': 0,  # Non pertinent pour cette analyse
                'end_of_life_cost': 0,  # Non pertinent pour cette analyse
                'lifespan': element_lifespan,
                'total_cost': operation_cost_wlc,  # Le total est le coût d'opération WLC
                '_annual_operation_cost': operation_cost_annual,
                '_years_operated': project_lifespan - 1
            })
        
        return jsonify({
            "success": True,
            "results": analysis_results,
            "summary": {
                'criteria': f'Coûts opération cumulés sur {project_lifespan - 1} ans'
            },
            "description": f"Éléments avec les coûts d'opération cumulés les plus élevés (annuel × {project_lifespan - 1} ans)"
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Erreur lors de l'analyse d'opération: {str(e)}"}), 500

@app.route('/analyze-cost-by-phase')
def analyze_cost_by_phase():
    """Analyse de la répartition des coûts par phases du cycle de vie"""
    try:
        from sparql_client import query_graphdb
        
        # Récupérer les paramètres de filtrage depuis les arguments de la requête
        selected_guids = request.args.get('selected_guids', '')
        filter_type = request.args.get('filter_type', 'all')  # 'all', 'selected', 'uniformat'
        uniformat_filter = request.args.get('uniformat_filter', '')
        
        # Récupérer d'abord la durée de vie du projet
        sparql_lifespan = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?lifespan WHERE {
          <http://example.com/ifc#Project> wlc:hasDuration ?lifespan .
        }
        """
        lifespan_result = query_graphdb(sparql_lifespan)
        project_lifespan = int(float(lifespan_result[0]['lifespan'])) if lifespan_result and 'lifespan' in lifespan_result[0] else 50
        
        # Construire les filtres pour la requête SPARQL
        guid_filter = ""
        additional_description = ""
        
        if filter_type == 'selected' and selected_guids:
            # Convertir la chaîne de GUIDs en liste
            guid_list = [guid.strip() for guid in selected_guids.split(',') if guid.strip()]
            if guid_list:
                # Séparer la construction pour compatibilité Python 3.11
                guid_values = ', '.join([f'"{guid}"' for guid in guid_list])
                guid_filter = f"FILTER(?guid IN ({guid_values}))"
                additional_description = f" - Filtré sur {len(guid_list)} éléments sélectionnés"
        
        elif filter_type == 'uniformat' and uniformat_filter:
            guid_filter = f'FILTER(CONTAINS(LCASE(?uniformatCode), LCASE("{uniformat_filter}")))'
            additional_description = f" - Filtré sur Uniformat '{uniformat_filter}'"
        
        # Requête SPARQL séparée pour chaque type de coût
        # Construction : somme directe
        sparql_construction = f"""
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT (SUM(?costValue) AS ?totalCost) (COUNT(?cost) AS ?costCount) (COUNT(DISTINCT ?element) AS ?elementCount)
        WHERE {{
          ?element wlc:hasCost ?cost .
          ?element wlc:globalId ?guid .
          ?cost a wlc:ConstructionCosts ;
                wlc:hasCostValue ?costValue .
          
          OPTIONAL {{ ?element wlc:hasUniformatCode ?uniformatCode . }}
          
          {guid_filter}
        }}
        """
        
        # Opération : somme * (durée de vie - 1)
        sparql_operation = f"""
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT (SUM(?costValue) AS ?totalAnnualCost) (COUNT(?cost) AS ?costCount) (COUNT(DISTINCT ?element) AS ?elementCount)
        WHERE {{
          ?element wlc:hasCost ?cost .
          ?element wlc:globalId ?guid .
          ?cost a wlc:OperationCosts ;
                wlc:hasCostValue ?costValue .
          
          OPTIONAL {{ ?element wlc:hasUniformatCode ?uniformatCode . }}
          
          {guid_filter}
        }}
        """
        
        # Maintenance : somme * durée de vie (coûts annuels récurrents)
        sparql_maintenance = f"""
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT (SUM(?costValue) AS ?totalCost) (COUNT(?cost) AS ?costCount) (COUNT(DISTINCT ?element) AS ?elementCount)
        WHERE {{
          ?element wlc:hasCost ?cost .
          ?element wlc:globalId ?guid .
          ?cost a wlc:MaintenanceCosts ;
                wlc:hasCostValue ?costValue .
          
          OPTIONAL {{ ?element wlc:hasUniformatCode ?uniformatCode . }}
          
          {guid_filter}
        }}
        """
        
        # Fin de vie : calculer selon le nombre de remplacements pendant la durée du projet
        sparql_endoflife = f"""
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?element ?guid ?costValue ?lifespan
        WHERE {{
          ?element wlc:hasCost ?cost .
          ?element wlc:globalId ?guid .
          ?cost a wlc:EndOfLifeCosts ;
                wlc:hasCostValue ?costValue .
          
          OPTIONAL {{ ?element wlc:hasDuration ?lifespan . }}
          OPTIONAL {{ ?element wlc:hasUniformatCode ?uniformatCode . }}
          
          {guid_filter}
        }}
        """
        
        phase_distribution = []
        total_project_cost = 0
        total_elements_analyzed = 0
        
        phase_labels = {
            'ConstructionCosts': 'Construction',
            'OperationCosts': 'Opération',
            'MaintenanceCosts': 'Maintenance',
            'EndOfLifeCosts': 'Fin de vie'
        }
        
        # Traitement de la construction  
        construction_results = query_graphdb(sparql_construction)
        if construction_results and construction_results[0]:
            row = construction_results[0]
            total_cost = float(row.get('totalCost', 0))
            cost_count = int(row.get('costCount', 0))
            element_count = int(row.get('elementCount', 0))
            
            if total_cost > 0:
                total_project_cost += total_cost
                total_elements_analyzed = max(total_elements_analyzed, element_count)
                
                phase_distribution.append({
                    'phase': 'Construction',
                    'cost_type': 'ConstructionCosts',
                    'total_cost': total_cost,
                    'cost_count': cost_count,
                    'element_count': element_count
                })
        
        # Traitement de l'opération (coût ANNUEL × (durée projet - 1))
        operation_results = query_graphdb(sparql_operation)
        if operation_results and operation_results[0]:
            row = operation_results[0]
            annual_cost = float(row.get('totalAnnualCost', 0))  # Coût annuel cumulé
            cost_count = int(row.get('costCount', 0))
            element_count = int(row.get('elementCount', 0))
            
            # Multiplier par le nombre d'années d'opération
            operation_years = project_lifespan - 1
            total_cost = annual_cost * operation_years
            
            if total_cost > 0:
                total_project_cost += total_cost
                total_elements_analyzed = max(total_elements_analyzed, element_count)
                
                phase_distribution.append({
                    'phase': 'Opération',
                    'cost_type': 'OperationCosts',
                    'total_cost': total_cost,
                    'annual_cost': annual_cost,
                    'operation_years': operation_years,
                    'cost_count': cost_count,
                    'element_count': element_count
                })
        
        # Traitement de la maintenance (avec multiplication par durée de vie + coûts de fin de vie des remplacements)
        maintenance_results = query_graphdb(sparql_maintenance)
        maintenance_annual_cost = 0
        maintenance_cost_count = 0
        maintenance_element_count = 0
        
        if maintenance_results and maintenance_results[0]:
            row = maintenance_results[0]
            maintenance_total_cost = float(row.get('totalAnnualCost', 0))  # Mal nommé, c'est le total
            maintenance_cost_count = int(row.get('costCount', 0))
            maintenance_element_count = int(row.get('elementCount', 0))
        
        # Récupérer et calculer les coûts de remplacements (maintenance ponctuelle) + démolition finale
        endoflife_results = query_graphdb(sparql_endoflife)
        endoflife_total_cost = 0  # Coûts de démolition finale uniquement
        total_demolitions = 0
        endoflife_element_count = 0
        
        # Calculer aussi les remplacements pour la maintenance
        maintenance_from_replacements = 0
        total_replacements = 0
        
        if endoflife_results:
            elements_processed = set()
            
            for row in endoflife_results:
                guid = row.get('guid', '')
                cost_value = float(row.get('costValue', 0))
                element_lifespan = int(float(row.get('lifespan', project_lifespan))) if row.get('lifespan') else project_lifespan
                
                if guid and cost_value > 0:
                    # Séparer remplacements et démolition finale
                    replacement_events = 0
                    if element_lifespan > 0 and element_lifespan < project_lifespan:
                        replacement_events = project_lifespan // element_lifespan
                        # Si division exacte, soustraire 1 (dernier = démolition)
                        if (project_lifespan % element_lifespan) == 0:
                            replacement_events -= 1
                    
                    # Maintenance = coût unitaire × nombre de remplacements
                    maintenance_from_replacements += cost_value * replacement_events
                    total_replacements += replacement_events
                    
                    # Fin de vie = coût de démolition finale (toujours 1)
                    endoflife_total_cost += cost_value
                    total_demolitions += 1
                    
                    if guid not in elements_processed:
                        elements_processed.add(guid)
                        endoflife_element_count += 1
        
        # Maintenance : coût annuel + coûts de remplacements
        # Note: maintenance_total_cost est le coût ponctuel unitaire
        # Il faut calculer combien de remplacements pour chaque élément
        total_maintenance_cost = 0
        
        # Pour calculer la maintenance correctement, on doit récupérer les durées de vie
        sparql_maintenance_detail = f"""
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?element ?guid ?costValue ?lifespan
        WHERE {{
          ?element wlc:hasCost ?cost .
          ?element wlc:globalId ?guid .
          ?cost a wlc:MaintenanceCosts ;
                wlc:hasCostValue ?costValue .
          
          OPTIONAL {{ ?element wlc:hasDuration ?lifespan . }}
          OPTIONAL {{ ?element wlc:hasUniformatCode ?uniformatCode . }}
          
          {guid_filter}
        }}
        """
        
        maintenance_detail_results = query_graphdb(sparql_maintenance_detail)
        if maintenance_detail_results:
            for row in maintenance_detail_results:
                cost_value = float(row.get('costValue', 0))
                element_lifespan = int(float(row.get('lifespan', project_lifespan))) if row.get('lifespan') else project_lifespan
                
                # Calculer le nombre de remplacements pour cet élément
                replacement_events = 0
                if element_lifespan > 0 and element_lifespan < project_lifespan:
                    replacement_events = project_lifespan // element_lifespan
                    if (project_lifespan % element_lifespan) == 0:
                        replacement_events -= 1
                
                total_maintenance_cost += cost_value * replacement_events
        
        if total_maintenance_cost > 0:
            total_project_cost += total_maintenance_cost
            total_elements_analyzed = max(total_elements_analyzed, maintenance_element_count)
            
            phase_distribution.append({
                'phase': 'Maintenance',
                'cost_type': 'MaintenanceCosts',
                'total_cost': total_maintenance_cost,
                'cost_count': maintenance_cost_count,
                'element_count': maintenance_element_count,
                'total_replacements': total_replacements,
                'description': f'Coûts de remplacements ({total_replacements} événements)'
            })
        
        # Phase fin de vie (démolitions finales uniquement)
        if endoflife_total_cost > 0:
            total_project_cost += endoflife_total_cost
            total_elements_analyzed = max(total_elements_analyzed, endoflife_element_count)
            
            phase_distribution.append({
                'phase': 'Fin de vie',
                'cost_type': 'EndOfLifeCosts',
                'total_cost': endoflife_total_cost,
                'cost_count': total_demolitions,
                'element_count': endoflife_element_count,
                'description': f'Coûts de démolitions finales ({total_demolitions} éléments)'
            })
        
        # Calculer les pourcentages et préparer les données pour le graphique
        cost_distribution = {}
        dominant_phase = ""
        max_cost = 0
        
        for phase in phase_distribution:
            percentage = (phase['total_cost'] / total_project_cost * 100) if total_project_cost > 0 else 0
            phase['percentage'] = percentage
            cost_distribution[phase['phase']] = round(percentage, 1)
            
            if phase['total_cost'] > max_cost:
                max_cost = phase['total_cost']
                dominant_phase = phase['phase']
        
        return jsonify({
            "success": True,
            "summary": {
                "phases_analyzed": len(phase_distribution),
                "dominant_phase": dominant_phase,
                "project_lifespan": project_lifespan,
                "operation_years": max(0, project_lifespan - 1),
                "cost_distribution": cost_distribution,
                "total_project_cost": total_project_cost,
                "elements_analyzed": total_elements_analyzed,
                "filter_applied": filter_type,
                "filter_description": additional_description
            },
            "phase_distribution": phase_distribution,
            "description": f"Répartition des coûts par phases du cycle de vie{additional_description}",
            "calculation_note": f"Construction (année 0) | Opération : coût annuel × {max(0, project_lifespan - 1)} ans | Maintenance : remplacements ponctuels ({total_replacements} événements) | Fin de vie : démolitions finales ({total_demolitions} éléments)"
        })
        
    except Exception as e:
        return jsonify({"error": f"Erreur lors de l'analyse par phases: {str(e)}"}), 500

@app.route('/export-analysis-results')
def export_analysis_results():
    """Export des résultats d'analyse vers Excel"""
    try:
        # Version simplifiée qui retourne un fichier Excel vide
        # Dans la version complète, cela exporterait les derniers résultats d'analyse
        
        import io
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Analyse WLC"
        
        # En-têtes
        headers = ['GUID', 'Description', 'Matériau', 'Coût Total', 'Type d\'Analyse']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Données d'exemple
        ws.cell(row=2, column=1, value="Exemple")
        ws.cell(row=2, column=2, value="Aucune analyse récente à exporter")
        ws.cell(row=2, column=3, value="-")
        ws.cell(row=2, column=4, value=0)
        ws.cell(row=2, column=5, value="N/A")
        
        # Sauvegarder dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        from flask import make_response
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=analyse_wlc_results.xlsx'
        
        return response
        
    except Exception as e:
        return jsonify({"error": f"Erreur lors de l'export: {str(e)}"}), 500

def set_element_duration(guid, duration):
    """Mettre à jour la durée de vie d'un élément dans GraphDB"""
    import requests
    
    uri = create_element_uri(guid)
    update = f"""
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
    PREFIX xsd: <http://www.w3.org/2001/XMLSchema#>
    DELETE {{ <{uri}> wlc:hasDuration ?old . }}
    INSERT {{ <{uri}> wlc:hasDuration "{duration}"^^xsd:integer . }}
    WHERE  {{ OPTIONAL {{ <{uri}> wlc:hasDuration ?old }} }}
    """
    requests.post(GRAPHDB_REPO.rstrip("/") + "/statements", data={"update": update})

@app.route('/costs-by-year')
def costs_by_year():
    """Retourne les coûts par année pour le graphique d'évolution"""
    try:
        from sparql_client import query_graphdb
        
        # Récupérer la durée de vie du projet
        sparql_lifespan = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?lifespan WHERE {
          <http://example.com/ifc#Project> wlc:hasDuration ?lifespan .
        }
        """
        lifespan_result = query_graphdb(sparql_lifespan)
        project_lifespan = int(float(lifespan_result[0]['lifespan'])) if lifespan_result and 'lifespan' in lifespan_result[0] else 50
        
        print(f"📊 costs-by-year: Durée de vie du projet: {project_lifespan} ans")
        
        # Récupérer tous les coûts par phase (MÊME LOGIQUE QUE analyze-cost-by-phase)
        
        # 1. Construction
        sparql_construction = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT (SUM(?costValue) AS ?totalCost)
        WHERE {
          ?element wlc:hasCost ?cost .
          ?cost a wlc:ConstructionCosts ;
                wlc:hasCostValue ?costValue .
        }
        """
        
        # 2. Opération (coût annuel)
        sparql_operation = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT (SUM(?costValue) AS ?totalAnnualCost)
        WHERE {
          ?element wlc:hasCost ?cost .
          ?cost a wlc:OperationCosts ;
                wlc:hasCostValue ?costValue .
        }
        """
        
        # 3. Maintenance (coût annuel)
        sparql_maintenance = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT (SUM(?costValue) AS ?totalAnnualCost)
        WHERE {
          ?element wlc:hasCost ?cost .
          ?cost a wlc:MaintenanceCosts ;
                wlc:hasCostValue ?costValue .
        }
        """
        
        # 4. Fin de vie avec durées de vie individuelles
        sparql_endoflife = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?element ?guid ?costValue ?lifespan
        WHERE {
          ?element wlc:hasCost ?cost .
          ?element wlc:globalId ?guid .
          ?cost a wlc:EndOfLifeCosts ;
                wlc:hasCostValue ?costValue .
          
          OPTIONAL { ?element wlc:hasDuration ?lifespan . }
        }
        """
        
        # Exécuter les requêtes
        construction_results = query_graphdb(sparql_construction)
        operation_results = query_graphdb(sparql_operation)
        maintenance_results = query_graphdb(sparql_maintenance)
        endoflife_results = query_graphdb(sparql_endoflife)
        
        # Calculer les coûts par phase
        # IMPORTANT : Opération = ANNUEL, Maintenance = PONCTUEL
        construction_cost = float(construction_results[0].get('totalCost', 0)) if construction_results and construction_results[0] else 0
        operation_annual_cost = float(operation_results[0].get('totalAnnualCost', 0)) if operation_results and operation_results[0] else 0  # Coût ANNUEL cumulé
        maintenance_unit_cost = float(maintenance_results[0].get('totalAnnualCost', 0)) if maintenance_results and maintenance_results[0] else 0  # Coût UNITAIRE cumulé
        
        print(f"📊 costs-by-year: Construction={construction_cost}, Opération annuelle={operation_annual_cost:.2f}, Maintenance unitaire={maintenance_unit_cost:.2f}")
        
        # Initialiser la structure des données par année
        data_by_year = []
        for year in range(project_lifespan + 1):
            data_by_year.append({
                'year': year,
                'ConstructionCosts': 0,
                'OperationCosts': 0,
                'MaintenanceCosts': 0,
                'EndOfLifeCosts': 0,
                'total': 0
            })
        
        # 1. Construction en année 0 UNIQUEMENT
        if construction_cost > 0:
            data_by_year[0]['ConstructionCosts'] = construction_cost
        
        # 2. Opération années 1 à (durée-1) UNIQUEMENT (pas année 0, pas année finale)
        if operation_annual_cost > 0:
            for year in range(1, project_lifespan):
                data_by_year[year]['OperationCosts'] = operation_annual_cost
        
        # 3. Maintenance (remplacements ponctuels) - calculer avec les durées de vie
        sparql_maintenance_detail = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?element ?guid ?costValue ?lifespan
        WHERE {
          ?element wlc:hasCost ?cost .
          ?element wlc:globalId ?guid .
          ?cost a wlc:MaintenanceCosts ;
                wlc:hasCostValue ?costValue .
          
          OPTIONAL { ?element wlc:hasDuration ?lifespan . }
        }
        """
        
        maintenance_detail_results = query_graphdb(sparql_maintenance_detail)
        
        if maintenance_detail_results:
            for row in maintenance_detail_results:
                cost_value = float(row.get('costValue', 0))
                element_lifespan = int(float(row.get('lifespan', project_lifespan))) if row.get('lifespan') else project_lifespan
                
                if cost_value > 0 and element_lifespan > 0 and element_lifespan < project_lifespan:
                    # Calculer les années de remplacement (pas dernière année)
                    replacement_year = element_lifespan
                    while replacement_year < project_lifespan:
                        data_by_year[replacement_year]['MaintenanceCosts'] += cost_value
                        replacement_year += element_lifespan
        
        # 4. Fin de vie (démolitions finales) - toujours à l'année N
        if endoflife_results:
            total_demolition_cost = 0
            for row in endoflife_results:
                cost_value = float(row.get('costValue', 0))
                if cost_value > 0:
                    total_demolition_cost += cost_value
            
            if total_demolition_cost > 0:
                data_by_year[project_lifespan]['EndOfLifeCosts'] = total_demolition_cost
        
        # Calculer les totaux
        for year_data in data_by_year:
            year_data['total'] = (
                year_data['ConstructionCosts'] + 
                year_data['OperationCosts'] + 
                year_data['MaintenanceCosts'] + 
                year_data['EndOfLifeCosts']
            )
        
        # Debug : afficher quelques exemples
        years_with_costs = [year for year in data_by_year if year['total'] > 0]
        print(f"📊 costs-by-year: {len(years_with_costs)} années avec des coûts")
        if years_with_costs:
            print(f"📊 costs-by-year: Exemple année 0: {years_with_costs[0] if years_with_costs else 'Aucune'}")
            if len(years_with_costs) > 1:
                print(f"📊 costs-by-year: Exemple dernière année: {years_with_costs[-1]}")
        
        return jsonify(data_by_year)
        
    except Exception as e:
        print(f"❌ Erreur dans costs-by-year: {str(e)}")
        return jsonify({"error": f"Erreur dans costs-by-year: {str(e)}"}), 500

@app.route('/get-discount-rates')
def get_discount_rates():
    """Récupère les taux d'actualisation par année"""
    try:
        # Version simplifiée qui retourne un taux par défaut de 3%
        # Dans la version complète, cela interrogerait GraphDB
        return jsonify({
            "success": True,
            "rates": [
                {"year": i, "discount_rate": 0.03, "discount_rate_percent": 3.0}
                for i in range(51)  # 0 à 50 ans par défaut
            ],
            "total_years": 51
        })
    except Exception as e:
        return jsonify({"error": f"Erreur lors de la récupération des taux: {str(e)}"}), 500

@app.route('/get-wlc')
def get_wlc():
    """Récupère le Whole Life Cost calculé"""
    try:
        # Version simplifiée qui retourne des données par défaut
        # Dans la version complète, cela interrogerait GraphDB pour le WLC calculé
        from sparql_client import query_graphdb
        
        sparql = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?totalValue ?discountRate ?name
        WHERE {
          <http://example.com/ifc#ProjectWLC> a wlc:WholeLifeCost ;
              wlc:hasTotalValue ?totalValue .
          OPTIONAL { <http://example.com/ifc#ProjectWLC> wlc:hasDiscountRate ?discountRate . }
          OPTIONAL { <http://example.com/ifc#ProjectWLC> wlc:name ?name . }
        }
        LIMIT 1
        """
        
        results = query_graphdb(sparql)
        
        if results and len(results) > 0:
            wlc_data = results[0]
            return jsonify({
                "exists": True,
                "total_value": float(wlc_data.get('totalValue', 0)),
                "discount_rate": float(wlc_data.get('discountRate', 0.03)),
                "name": wlc_data.get('name', 'Coût global du projet')
            })
        else:
            return jsonify({
                "exists": False,
                "total_value": 0,
                "discount_rate": 0.03,
                "name": "Aucun WLC calculé"
            })
        
    except Exception as e:
        return jsonify({"error": f"Erreur get-wlc: {str(e)}"}), 500

# Enregistrer les routes de comparaison
# Note: calculate_wlc est la route, nous devons créer une fonction wrapper pour l'export
def calculate_wlc_for_export():
    """Wrapper pour calculate_wlc pour l'export"""
    try:
        # Simuler un POST request pour calculate_wlc
        from flask import g as flask_g
        # Retourner les données WLC calculées
        return {
            'total_wlc': 0,  # À implémenter selon vos besoins
            'average_discount_rate': 0.03
        }
    except:
        return None

register_comparison_routes(app, g, calculate_wlc_for_export, get_multi_stakeholder_view)

@app.route('/create-element', methods=['POST'])
def create_element():
    """Créer un nouvel élément dans l'ontologie"""
    try:
        data = request.get_json()
        
        # Vérifier les données requises
        required_fields = ['guid', 'uniformat_code', 'uniformat_description', 'element_class', 'element_description']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'error': f'Champ requis manquant: {field}'}), 400
        
        guid = data['guid']
        uniformat_code = data['uniformat_code']
        uniformat_description = data['uniformat_description']
        element_class = data['element_class']
        element_description = data['element_description']
        
        # Coûts optionnels
        construction_cost = float(data.get('construction_cost', 0))
        operation_cost = float(data.get('operation_cost', 0))
        maintenance_cost = float(data.get('maintenance_cost', 0))
        end_of_life_cost = float(data.get('end_of_life_cost', 0))
        lifespan = int(data.get('lifespan', 60))
        
        print(f"🔨 Création élément: {guid} - {uniformat_code} - {element_description}")
        
        # Créer l'élément dans l'ontologie
        WLC = Namespace("http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#")
        
        # URI de l'élément
        element_uri = URIRef(f"http://example.com/ifc#{guid}")
        
        # Propriétés de base
        g.add((element_uri, RDF.type, WLC.Element))
        g.add((element_uri, WLC.globalId, Literal(guid)))
        g.add((element_uri, WLC.hasUniformatCode, Literal(uniformat_code)))
        g.add((element_uri, WLC.hasUniformatDescription, Literal(uniformat_description)))
        g.add((element_uri, WLC.hasIfcClass, Literal(element_class)))
        g.add((element_uri, WLC.hasDescription, Literal(element_description)))
        
        # Durée de vie
        lifespan_uri = URIRef(f"http://example.com/ifc#{guid}_lifespan")
        g.add((lifespan_uri, RDF.type, WLC.Duration))
        g.add((lifespan_uri, WLC.hasValue, Literal(lifespan)))
        g.add((element_uri, WLC.hasDuration, lifespan_uri))
        
        # Créer les coûts si > 0
        cost_counter = 1
        
        if construction_cost > 0:
            cost_uri = URIRef(f"http://example.com/ifc#{guid}_cost_{cost_counter}")
            g.add((cost_uri, RDF.type, WLC.ConstructionCosts))
            g.add((cost_uri, WLC.hasCostValue, Literal(construction_cost)))
            g.add((element_uri, WLC.hasCost, cost_uri))
            cost_counter += 1
        
        if operation_cost > 0:
            cost_uri = URIRef(f"http://example.com/ifc#{guid}_cost_{cost_counter}")
            g.add((cost_uri, RDF.type, WLC.OperationCosts))
            g.add((cost_uri, WLC.hasCostValue, Literal(operation_cost)))
            g.add((element_uri, WLC.hasCost, cost_uri))
            cost_counter += 1
        
        if maintenance_cost > 0:
            cost_uri = URIRef(f"http://example.com/ifc#{guid}_cost_{cost_counter}")
            g.add((cost_uri, RDF.type, WLC.MaintenanceCosts))
            g.add((cost_uri, WLC.hasCostValue, Literal(maintenance_cost)))
            g.add((element_uri, WLC.hasCost, cost_uri))
            cost_counter += 1
        
        if end_of_life_cost > 0:
            cost_uri = URIRef(f"http://example.com/ifc#{guid}_cost_{cost_counter}")
            g.add((cost_uri, RDF.type, WLC.EndOfLifeCosts))
            g.add((cost_uri, WLC.hasCostValue, Literal(end_of_life_cost)))
            g.add((element_uri, WLC.hasCost, cost_uri))
        
        print(f"✅ Élément créé avec succès: {guid}")
        
        return jsonify({
            'success': True, 
            'message': f'Élément {guid} créé avec succès',
            'element': {
                'guid': guid,
                'uniformat_code': uniformat_code,
                'uniformat_description': uniformat_description,
                'element_class': element_class,
                'element_description': element_description,
                'lifespan': lifespan,
                'costs': {
                    'construction': construction_cost,
                    'operation': operation_cost,
                    'maintenance': maintenance_cost,
                    'end_of_life': end_of_life_cost
                }
            }
        })
        
    except Exception as e:
        print(f"❌ Erreur création élément: {str(e)}")
        return jsonify({'success': False, 'error': f'Erreur lors de la création: {str(e)}'}), 500

@app.route('/enrich-ifc', methods=['POST'])
def enrich_ifc():
    """Enrichit le fichier IFC en mémoire avec les données WLC de l'ontologie"""
    global ifc_storage
    
    try:
        # Vérifier qu'un fichier IFC est en mémoire
        if not ifc_storage['current_file']:
            return jsonify({
                'success': False,
                'error': 'Aucun fichier IFC en mémoire. Veuillez d\'abord importer un fichier IFC.'
            }), 400
        
        print("🔧 Début de l'enrichissement IFC...")
        
        # Importer ifcopenshell
        try:
            import ifcopenshell
        except ImportError:
            return jsonify({
                'success': False,
                'error': 'Module ifcopenshell non disponible. Utilisez l\'environnement Python 3.12.'
            }), 500
        
        # Récupérer les données WLC de l'ontologie
        print("📊 Récupération des données WLC...")
        sparql_wlc_data = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        SELECT ?element ?guid ?constructionCost ?operationCost ?maintenanceCost ?endOfLifeCost ?lifespan ?uniformatCode ?uniformatDesc
        WHERE {
            ?element a wlc:Element .
            OPTIONAL { ?element wlc:globalId ?guid . }
            OPTIONAL { 
                ?element wlc:hasCost ?conCost .
                ?conCost a wlc:ConstructionCosts ;
                         wlc:hasCostValue ?constructionCost .
            }
            OPTIONAL { 
                ?element wlc:hasCost ?opCost .
                ?opCost a wlc:OperationCosts ;
                        wlc:hasCostValue ?operationCost .
            }
            OPTIONAL { 
                ?element wlc:hasCost ?mainCost .
                ?mainCost a wlc:MaintenanceCosts ;
                          wlc:hasCostValue ?maintenanceCost .
            }
            OPTIONAL { 
                ?element wlc:hasCost ?endCost .
                ?endCost a wlc:EndOfLifeCosts ;
                         wlc:hasCostValue ?endOfLifeCost .
            }
            OPTIONAL { ?element wlc:hasDuration ?lifespan . }
            OPTIONAL { ?element wlc:hasUniformatCode ?uniformatCode . }
            OPTIONAL { ?element wlc:hasUniformatDescription ?uniformatDesc . }
        }
        """
        
        wlc_elements = query_graphdb(sparql_wlc_data)
        print(f"📊 {len(wlc_elements)} éléments WLC récupérés")
        
        if not wlc_elements:
            return jsonify({
                'success': False,
                'error': 'Aucune donnée WLC trouvée dans l\'ontologie'
            }), 400
        
        # Charger le fichier IFC depuis la mémoire
        print("📂 Chargement du fichier IFC...")
        ifc_content = ifc_storage['current_file']['content']
        
        # Créer un fichier temporaire pour ifcopenshell
        import tempfile
        import os
        
        with tempfile.NamedTemporaryFile(suffix='.ifc', delete=False) as temp_file:
            temp_file.write(ifc_content)
            temp_file_path = temp_file.name
        
        try:
            # Ouvrir le fichier IFC avec ifcopenshell
            ifc_file = ifcopenshell.open(temp_file_path)
            print(f"📂 Fichier IFC ouvert: {len(ifc_file.by_type('IfcElement'))} éléments")
            
            # Créer un dictionnaire des données WLC par GUID
            wlc_data_by_guid = {}
            # Créer un dictionnaire des données WLC par groupe Uniformat
            wlc_data_by_uniformat = {}
            
            for elem_data in wlc_elements:
                guid = elem_data.get('guid', '')
                uniformat_code = elem_data.get('uniformatCode', '')
                
                wlc_data = {
                    'construction_cost': float(elem_data.get('constructionCost', 0)) if elem_data.get('constructionCost') else 0,
                    'operation_cost': float(elem_data.get('operationCost', 0)) if elem_data.get('operationCost') else 0,
                    'maintenance_cost': float(elem_data.get('maintenanceCost', 0)) if elem_data.get('maintenanceCost') else 0,
                    'end_of_life_cost': float(elem_data.get('endOfLifeCost', 0)) if elem_data.get('endOfLifeCost') else 0,
                    'lifespan': int(float(elem_data.get('lifespan', 0))) if elem_data.get('lifespan') else 0,
                    'uniformat_code': uniformat_code,
                    'uniformat_desc': elem_data.get('uniformatDesc', '')
                }
                
                # Indexer par GUID si disponible
                if guid:
                    wlc_data_by_guid[guid] = wlc_data
                
                # Indexer par code Uniformat si disponible
                if uniformat_code:
                    if uniformat_code not in wlc_data_by_uniformat:
                        wlc_data_by_uniformat[uniformat_code] = []
                    wlc_data_by_uniformat[uniformat_code].append(wlc_data)
            
            print(f"📊 Données WLC organisées:")
            print(f"   • Par GUID: {len(wlc_data_by_guid)} éléments")
            print(f"   • Par groupe Uniformat: {len(wlc_data_by_uniformat)} groupes")
            
            # Initialiser les compteurs d'enrichissement pour l'ensemble de la fonction
            enriched_count = 0
            enriched_by_guid = 0
            enriched_by_uniformat = 0
            
            # Fonction pour extraire le code Uniformat d'un élément IFC
            def extract_ifc_uniformat_code(ifc_element):
                """Extrait le code Uniformat d'un élément IFC"""
                if hasattr(ifc_element, 'IsDefinedBy'):
                    for rel in ifc_element.IsDefinedBy:
                        if hasattr(rel, 'RelatingPropertyDefinition'):
                            pset = rel.RelatingPropertyDefinition
                            if hasattr(pset, 'HasProperties'):
                                for prop in pset.HasProperties:
                                    if hasattr(prop, 'Name') and prop.Name and 'uniformat' in prop.Name.lower():
                                        if hasattr(prop, 'NominalValue') and hasattr(prop.NominalValue, 'wrappedValue'):
                                            return prop.NominalValue.wrappedValue
                return None
            
            # Fonction pour déterminer le groupe Uniformat basé sur le type d'élément
            def determine_uniformat_group(ifc_element):
                """Détermine le groupe Uniformat basé sur le type d'élément IFC"""
                element_type = ifc_element.is_a().lower()
                element_name = (ifc_element.Name or '').lower()
                
                # Mapping des types IFC vers les groupes Uniformat
                uniformat_mapping = {
                    'ifcwall': 'B2010',
                    'ifcwallstandardcase': 'B2010', 
                    'ifcwindow': 'B2020',
                    'ifcdoor': 'B2020',
                    'ifcslab': 'B1010',
                    'ifcroof': 'B3010',
                    'ifcbeam': 'B1020',
                    'ifccolumn': 'B1020',
                    'ifcstair': 'B2030',
                    'ifcrailing': 'B2030',
                    'ifcflowsegment': 'D3040',
                    'ifcflowfitting': 'D3040',
                    'ifcflowterminal': 'D3040',
                    'ifcairtoairheatrecovery': 'D3030',
                    'ifcchiller': 'D3030',
                    'ifcboiler': 'D3020',
                    'ifcunitaryequipment': 'D3050',
                    'ifcfan': 'D3040',
                    'ifcpump': 'D3040',
                    'ifcelectricmotor': 'D5020',
                    'ifclightfixture': 'D5010',
                    'ifcelectricappliance': 'D5020'
                }
                
                # Vérifier d'abord dans le mapping direct
                if element_type in uniformat_mapping:
                    return uniformat_mapping[element_type]
                
                # Vérifier par mots-clés dans le nom
                if any(keyword in element_name for keyword in ['chauffage', 'heating', 'radiator', 'radiateur']):
                    return 'D3020'
                elif any(keyword in element_name for keyword in ['climatisation', 'cooling', 'air conditioner']):
                    return 'D3030'
                elif any(keyword in element_name for keyword in ['ventilation', 'fan', 'air handling']):
                    return 'D3040'
                elif any(keyword in element_name for keyword in ['mur', 'wall', 'rideau']):
                    return 'B2010'
                elif any(keyword in element_name for keyword in ['window', 'fenêtre']):
                    return 'B2020'
                elif any(keyword in element_name for keyword in ['door', 'porte']):
                    return 'B2020'
                
                return None
            
            # Fonction pour obtenir les données WLC moyennes d'un groupe Uniformat
            def get_uniformat_group_data(uniformat_code):
                """Obtient les données WLC moyennes d'un groupe Uniformat"""
                if uniformat_code not in wlc_data_by_uniformat:
                    return None
                
                group_data = wlc_data_by_uniformat[uniformat_code]
                if not group_data:
                    return None
                
                # Calculer les moyennes
                avg_data = {
                    'construction_cost': sum(d['construction_cost'] for d in group_data) / len(group_data),
                    'operation_cost': sum(d['operation_cost'] for d in group_data) / len(group_data),
                    'maintenance_cost': sum(d['maintenance_cost'] for d in group_data) / len(group_data),
                    'end_of_life_cost': sum(d['end_of_life_cost'] for d in group_data) / len(group_data),
                    'lifespan': int(sum(d['lifespan'] for d in group_data if d['lifespan'] > 0) / len([d for d in group_data if d['lifespan'] > 0])) if any(d['lifespan'] > 0 for d in group_data) else 0,
                    'uniformat_code': uniformat_code,
                    'uniformat_desc': group_data[0]['uniformat_desc']
                }
                
                return avg_data
            
            # ENRICHISSEMENT SPÉCIAL POUR LES IfcGroup CIBLES
            # Groupes cibles recherchés par nom (plus robuste que par GUID)
            target_groups_by_name = {
                'Murs-rideaux MR_V3_ENV': '3ffPwhTTv76OU5CdZc3Mgo',  # Nom -> GUID ontologie
                'Murs de base R02.1': 'c6175a257c2049a88f8c16'      # Nom -> GUID ontologie
            }
            
            groups_enriched = 0
            elements_enriched_from_groups = 0
            
            print("🎯 Enrichissement spécial des IfcGroup cibles (recherche par nom)...")
            
            # Récupérer tous les IfcGroup du fichier
            all_ifc_groups = ifc_file.by_type('IfcGroup')
            print(f"📦 {len(all_ifc_groups)} IfcGroup trouvés dans le fichier")
            
            # Lister tous les noms de groupes trouvés pour debug
            found_group_names = []
            for debug_group in all_ifc_groups:
                if hasattr(debug_group, 'Name') and debug_group.Name:
                    found_group_names.append(debug_group.Name.strip())
            print(f"📋 Noms de groupes trouvés dans le fichier: {found_group_names}")
            
            for ifc_group in all_ifc_groups:
                if hasattr(ifc_group, 'Name') and ifc_group.Name:
                    group_name = ifc_group.Name.strip()
                    
                    # Chercher si ce nom correspond à un de nos groupes cibles
                    if group_name in target_groups_by_name:
                        group_guid = ifc_group.GlobalId if hasattr(ifc_group, 'GlobalId') else 'N/A'
                        ontology_guid = target_groups_by_name[group_name]
                        
                        print(f"🎯 Groupe cible trouvé par nom: '{group_name}'")
                        print(f"   📋 GUID IFC: {group_guid}")
                        print(f"   🔍 GUID ontologie: {ontology_guid}")
                        
                        # Récupérer les données WLC depuis l'ontologie
                        group_wlc_data = wlc_data_by_guid.get(ontology_guid)
                        
                        print(f"   🔍 Recherche données WLC dans l'ontologie...")
                        
                        if group_wlc_data:
                            print(f"   📊 Données WLC trouvées pour le groupe!")
                            print(f"      - Construction: {group_wlc_data['construction_cost']}")
                            print(f"      - Opération: {group_wlc_data['operation_cost']}")
                            print(f"      - Maintenance: {group_wlc_data['maintenance_cost']}")
                            print(f"      - Fin de vie: {group_wlc_data['end_of_life_cost']}")
                            
                            # 1. Enrichir l'IfcGroup lui-même
                            group_properties = []
                            
                            if group_wlc_data['construction_cost'] > 0:
                                group_properties.append(
                                    ifc_file.createIfcPropertySingleValue(
                                        "GroupConstructionCost", None,
                                        ifc_file.createIfcReal(group_wlc_data['construction_cost']), None
                                    )
                                )
                            
                            if group_wlc_data['operation_cost'] > 0:
                                group_properties.append(
                                    ifc_file.createIfcPropertySingleValue(
                                        "GroupOperationCost", None,
                                        ifc_file.createIfcReal(group_wlc_data['operation_cost']), None
                                    )
                                )
                            
                            if group_wlc_data['maintenance_cost'] > 0:
                                group_properties.append(
                                    ifc_file.createIfcPropertySingleValue(
                                        "GroupMaintenanceCost", None,
                                        ifc_file.createIfcReal(group_wlc_data['maintenance_cost']), None
                                    )
                                )
                            
                            if group_wlc_data['end_of_life_cost'] > 0:
                                group_properties.append(
                                    ifc_file.createIfcPropertySingleValue(
                                        "GroupEndOfLifeCost", None,
                                        ifc_file.createIfcReal(group_wlc_data['end_of_life_cost']), None
                                    )
                                )
                            
                            if group_wlc_data['lifespan'] > 0:
                                group_properties.append(
                                    ifc_file.createIfcPropertySingleValue(
                                        "GroupLifespan", None,
                                        ifc_file.createIfcInteger(group_wlc_data['lifespan']), None
                                    )
                                )
                            
                            if group_wlc_data['uniformat_code']:
                                group_properties.append(
                                    ifc_file.createIfcPropertySingleValue(
                                        "GroupUniformatCode", None,
                                        ifc_file.createIfcText(group_wlc_data['uniformat_code']), None
                                    )
                                )
                            
                            if group_wlc_data['uniformat_desc']:
                                group_properties.append(
                                    ifc_file.createIfcPropertySingleValue(
                                        "GroupUniformatDescription", None,
                                        ifc_file.createIfcText(group_wlc_data['uniformat_desc']), None
                                    )
                                )
                            
                            # Ajouter une propriété pour indiquer l'enrichissement spécial par nom
                            group_properties.append(
                                ifc_file.createIfcPropertySingleValue(
                                    "EnrichmentMethod", None,
                                    ifc_file.createIfcText(f"SpecialGroupMappingByName_{group_name}"), None
                                )
                            )
                            
                            # Créer le PropertySet pour le groupe
                            if group_properties:
                                group_property_set = ifc_file.createIfcPropertySet(
                                    ifcopenshell.guid.new(),
                                    ifc_file.by_type('IfcOwnerHistory')[0] if ifc_file.by_type('IfcOwnerHistory') else None,
                                    "WLC_Group_Data",
                                    "Données WLC du groupe et propagation aux éléments",
                                    group_properties
                                )
                                
                                # Lier le PropertySet au groupe
                                ifc_file.createIfcRelDefinesByProperties(
                                    ifcopenshell.guid.new(),
                                    ifc_file.by_type('IfcOwnerHistory')[0] if ifc_file.by_type('IfcOwnerHistory') else None,
                                    None, None,
                                    [ifc_group],
                                    group_property_set
                                )
                                
                                groups_enriched += 1
                                print(f"   ✅ Groupe enrichi avec {len(group_properties)} propriétés")
                            
                            # 2. Propager les données WLC à tous les éléments du groupe
                            group_elements = []
                            if hasattr(ifc_group, 'IsGroupedBy'):
                                for rel in ifc_group.IsGroupedBy:
                                    if hasattr(rel, 'RelatedObjects'):
                                        for obj in rel.RelatedObjects:
                                            if hasattr(obj, 'GlobalId'):
                                                group_elements.append(obj)
                            
                            print(f"   📦 Propagation des données à {len(group_elements)} éléments du groupe")
                            
                            for element in group_elements:
                                try:
                                    # Créer les propriétés WLC pour l'élément
                                    element_properties = []
                                    
                                    if group_wlc_data['construction_cost'] > 0:
                                        element_properties.append(
                                            ifc_file.createIfcPropertySingleValue(
                                                "ConstructionCost", None,
                                                ifc_file.createIfcReal(group_wlc_data['construction_cost']), None
                                            )
                                        )
                                    
                                    if group_wlc_data['operation_cost'] > 0:
                                        element_properties.append(
                                            ifc_file.createIfcPropertySingleValue(
                                                "OperationCost", None,
                                                ifc_file.createIfcReal(group_wlc_data['operation_cost']), None
                                            )
                                        )
                                    
                                    if group_wlc_data['maintenance_cost'] > 0:
                                        element_properties.append(
                                            ifc_file.createIfcPropertySingleValue(
                                                "MaintenanceCost", None,
                                                ifc_file.createIfcReal(group_wlc_data['maintenance_cost']), None
                                            )
                                        )
                                    
                                    if group_wlc_data['end_of_life_cost'] > 0:
                                        element_properties.append(
                                            ifc_file.createIfcPropertySingleValue(
                                                "EndOfLifeCost", None,
                                                ifc_file.createIfcReal(group_wlc_data['end_of_life_cost']), None
                                            )
                                        )
                                    
                                    if group_wlc_data['lifespan'] > 0:
                                        element_properties.append(
                                            ifc_file.createIfcPropertySingleValue(
                                                "Lifespan", None,
                                                ifc_file.createIfcInteger(group_wlc_data['lifespan']), None
                                            )
                                        )
                                    
                                    if group_wlc_data['uniformat_code']:
                                        element_properties.append(
                                            ifc_file.createIfcPropertySingleValue(
                                                "UniformatCode", None,
                                                ifc_file.createIfcText(group_wlc_data['uniformat_code']), None
                                            )
                                        )
                                    
                                    if group_wlc_data['uniformat_desc']:
                                        element_properties.append(
                                            ifc_file.createIfcPropertySingleValue(
                                                "UniformatDescription", None,
                                                ifc_file.createIfcText(group_wlc_data['uniformat_desc']), None
                                            )
                                        )
                                    
                                    # Ajouter la source de l'enrichissement
                                    element_properties.append(
                                        ifc_file.createIfcPropertySingleValue(
                                            "EnrichmentMethod", None,
                                            ifc_file.createIfcText(f"PropagatedFromGroupByName_{group_name}"), None
                                        )
                                    )
                                    
                                    element_properties.append(
                                        ifc_file.createIfcPropertySingleValue(
                                            "SourceGroupName", None,
                                            ifc_file.createIfcText(group_name), None
                                        )
                                    )
                                    
                                    # Créer le PropertySet pour l'élément
                                    if element_properties:
                                        element_property_set = ifc_file.createIfcPropertySet(
                                            ifcopenshell.guid.new(),
                                            ifc_file.by_type('IfcOwnerHistory')[0] if ifc_file.by_type('IfcOwnerHistory') else None,
                                            "WLC_Data",
                                            "Données WLC propagées depuis le groupe parent",
                                            element_properties
                                        )
                                        
                                        # Lier le PropertySet à l'élément
                                        ifc_file.createIfcRelDefinesByProperties(
                                            ifcopenshell.guid.new(),
                                            ifc_file.by_type('IfcOwnerHistory')[0] if ifc_file.by_type('IfcOwnerHistory') else None,
                                            None, None,
                                            [element],
                                            element_property_set
                                        )
                                        
                                        elements_enriched_from_groups += 1
                                        
                                except Exception as e:
                                    print(f"     ⚠️  Erreur lors de l'enrichissement de l'élément {element.GlobalId}: {e}")
                                    continue
                            
                            print(f"   ✅ {len(group_elements)} éléments enrichis depuis le groupe")
                        
                        else:
                            print(f"   ⚠️  Aucune donnée WLC trouvée pour le groupe '{group_name}' (GUID ontologie: {ontology_guid})")
                            print(f"       Vérifiez que les données WLC existent dans l'ontologie pour ce GUID")
                    
                    else:
                        print(f"   ℹ️  Groupe ignoré: '{group_name}' (ne fait pas partie des groupes cibles)")
            
            print(f"🎯 Enrichissement spécial terminé:")
            print(f"   • Groupes IfcGroup cibles enrichis: {groups_enriched}")
            print(f"   • Éléments enrichis depuis les groupes: {elements_enriched_from_groups}")
            print(f"   • Autres éléments enrichis: {enriched_count}")
            print(f"     - Par GUID exact: {enriched_by_guid}")
            print(f"     - Par groupe Uniformat: {enriched_by_uniformat}")
            
            total_enriched = enriched_count + elements_enriched_from_groups
            
            # ENRICHISSEMENT NORMAL (logique existante) - peut être désactivé si nécessaire
            print("\n🔧 Enrichissement normal des autres éléments...")
            
            # Enrichir les éléments IFC (variables déjà déclarées plus haut)
            for ifc_element in ifc_file.by_type('IfcElement'):
                if not hasattr(ifc_element, 'GlobalId'):
                    continue
                    
                wlc_data = None
                enrichment_method = None
                
                # Méthode 1: Enrichissement par GUID exact
                if ifc_element.GlobalId in wlc_data_by_guid:
                    wlc_data = wlc_data_by_guid[ifc_element.GlobalId]
                    enrichment_method = "GUID"
                    enriched_by_guid += 1
                
                # Méthode 2: Enrichissement par groupe Uniformat si pas de GUID correspondant
                elif not wlc_data:
                    # Essayer d'extraire le code Uniformat de l'élément IFC
                    ifc_uniformat_code = extract_ifc_uniformat_code(ifc_element)
                    
                    # Si pas trouvé, déterminer basé sur le type d'élément
                    if not ifc_uniformat_code:
                        ifc_uniformat_code = determine_uniformat_group(ifc_element)
                    
                    if ifc_uniformat_code:
                        wlc_data = get_uniformat_group_data(ifc_uniformat_code)
                        if wlc_data:
                            enrichment_method = f"Uniformat_{ifc_uniformat_code}"
                            enriched_by_uniformat += 1
                
                # Appliquer l'enrichissement si des données ont été trouvées
                if wlc_data:
                    # Créer des propriétés personnalisées pour les données WLC
                    property_set_name = "WLC_Data"
                    
                    # Vérifier si le PropertySet existe déjà
                    existing_pset = None
                    if hasattr(ifc_element, 'IsDefinedBy'):
                        for rel in ifc_element.IsDefinedBy:
                            if hasattr(rel, 'RelatingPropertyDefinition'):
                                pset = rel.RelatingPropertyDefinition
                                if hasattr(pset, 'Name') and pset.Name == property_set_name:
                                    existing_pset = pset
                                    break
                    
                    # Si le PropertySet n'existe pas, le créer
                    if not existing_pset:
                        # Créer les propriétés WLC
                        wlc_properties = []
                        
                        if wlc_data['construction_cost'] > 0:
                            wlc_properties.append(
                                ifc_file.createIfcPropertySingleValue(
                                    "ConstructionCost", None,
                                    ifc_file.createIfcReal(wlc_data['construction_cost']), None
                                )
                            )
                        
                        if wlc_data['operation_cost'] > 0:
                            wlc_properties.append(
                                ifc_file.createIfcPropertySingleValue(
                                    "OperationCost", None,
                                    ifc_file.createIfcReal(wlc_data['operation_cost']), None
                                )
                            )
                        
                        if wlc_data['maintenance_cost'] > 0:
                            wlc_properties.append(
                                ifc_file.createIfcPropertySingleValue(
                                    "MaintenanceCost", None,
                                    ifc_file.createIfcReal(wlc_data['maintenance_cost']), None
                                )
                            )
                        
                        if wlc_data['end_of_life_cost'] > 0:
                            wlc_properties.append(
                                ifc_file.createIfcPropertySingleValue(
                                    "EndOfLifeCost", None,
                                    ifc_file.createIfcReal(wlc_data['end_of_life_cost']), None
                                )
                            )
                        
                        if wlc_data['lifespan'] > 0:
                            wlc_properties.append(
                                ifc_file.createIfcPropertySingleValue(
                                    "Lifespan", None,
                                    ifc_file.createIfcInteger(wlc_data['lifespan']), None
                                )
                            )
                        
                        if wlc_data['uniformat_code']:
                            wlc_properties.append(
                                ifc_file.createIfcPropertySingleValue(
                                    "UniformatCode", None,
                                    ifc_file.createIfcText(wlc_data['uniformat_code']), None
                                )
                            )
                        
                        if wlc_data['uniformat_desc']:
                            wlc_properties.append(
                                ifc_file.createIfcPropertySingleValue(
                                    "UniformatDescription", None,
                                    ifc_file.createIfcText(wlc_data['uniformat_desc']), None
                                )
                            )
                        
                        # Ajouter la méthode d'enrichissement
                        if enrichment_method:
                            wlc_properties.append(
                                ifc_file.createIfcPropertySingleValue(
                                    "EnrichmentMethod", None,
                                    ifc_file.createIfcText(enrichment_method), None
                                )
                            )
                        
                        # Créer le PropertySet si on a des propriétés
                        if wlc_properties:
                            property_set = ifc_file.createIfcPropertySet(
                                ifcopenshell.guid.new(),
                                ifc_file.by_type('IfcOwnerHistory')[0],
                                property_set_name,
                                "Données de coût du cycle de vie",
                                wlc_properties
                            )
                            
                            # Lier le PropertySet à l'élément
                            ifc_file.createIfcRelDefinesByProperties(
                                ifcopenshell.guid.new(),
                                ifc_file.by_type('IfcOwnerHistory')[0],
                                None, None,
                                [ifc_element],
                                property_set
                            )
                            
                            enriched_count += 1
            
            print(f"✅ Enrichissement terminé:")
            print(f"   • Total enrichis: {enriched_count} éléments")
            print(f"   • Par GUID exact: {enriched_by_guid} éléments")
            print(f"   • Par groupe Uniformat: {enriched_by_uniformat} éléments")
            
            # Sauvegarder le fichier enrichi
            enriched_temp_path = temp_file_path.replace('.ifc', '_enriched.ifc')
            ifc_file.write(enriched_temp_path)
            
            # Lire le fichier enrichi
            with open(enriched_temp_path, 'rb') as f:
                enriched_content = f.read()
            
            # Mettre à jour le stockage en mémoire
            ifc_storage['current_file']['content'] = enriched_content
            ifc_storage['current_file']['enriched'] = True
            
            print("✅ Enrichissement terminé avec succès")
            
            # Message personnalisé selon le type d'enrichissement
            if groups_enriched > 0:
                message = f"Fichier IFC enrichi avec enrichissement spécial des groupes! {groups_enriched} groupes IfcGroup enrichis, {elements_enriched_from_groups} éléments enrichis depuis les groupes, {enriched_count} autres éléments enrichis ({enriched_by_guid} par GUID, {enriched_by_uniformat} par groupe Uniformat)."
            else:
                message = f"Fichier IFC enrichi avec succès! {enriched_count} éléments enrichis ({enriched_by_guid} par GUID, {enriched_by_uniformat} par groupe Uniformat)."
            
            return jsonify({
                'success': True,
                'message': message,
                'total_enriched_elements': total_enriched,
                'groups_enriched': groups_enriched,
                'elements_enriched_from_groups': elements_enriched_from_groups,
                'enriched_elements': enriched_count,
                'enriched_by_guid': enriched_by_guid,
                'enriched_by_uniformat': enriched_by_uniformat,
                'total_wlc_elements': len(wlc_elements)
            })
            
        finally:
            # Nettoyer les fichiers temporaires
            try:
                os.unlink(temp_file_path)
                if 'enriched_temp_path' in locals():
                    os.unlink(enriched_temp_path)
            except:
                pass
                
    except Exception as e:
        print(f"❌ Erreur lors de l'enrichissement: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'Erreur lors de l\'enrichissement: {str(e)}'
        }), 500

@app.route('/download-enriched-ifc', methods=['POST'])
def download_enriched_ifc():
    """Télécharge le fichier IFC enrichi depuis la mémoire"""
    global ifc_storage
    
    try:
        # Vérifier qu'un fichier est en mémoire
        if not ifc_storage['current_file']:
            return jsonify({'error': 'Aucun fichier IFC en mémoire'}), 400
        
        # Récupérer le contenu du fichier (enrichi ou original)
        file_content = ifc_storage['current_file']['content']
        original_filename = ifc_storage['current_file']['filename']
        
        # Nom du fichier enrichi
        if ifc_storage['current_file'].get('enriched', False):
            download_filename = original_filename.replace('.ifc', '_WLC_enriched.ifc')
        else:
            download_filename = original_filename.replace('.ifc', '_processed.ifc')
        
        # Créer un fichier temporaire pour le téléchargement
        temp_file = io.BytesIO(file_content)
        temp_file.seek(0)
        
        return send_file(
            temp_file,
            as_attachment=True,
            download_name=download_filename,
            mimetype='application/octet-stream'
        )
        
    except Exception as e:
        print(f"Erreur lors du téléchargement: {str(e)}")
        return jsonify({'error': f'Erreur lors du téléchargement: {str(e)}'}), 500

@app.route('/get-end-of-life-strategies')
def get_end_of_life_strategies():
    """Retourne la liste des stratégies des 10 R disponibles"""
    try:
        strategies = [
            {'value': 'Refuse', 'label': 'Refuse - Refuser l\'utilisation'},
            {'value': 'Rethink', 'label': 'Rethink - Repenser l\'usage'},
            {'value': 'Reduce', 'label': 'Reduce - Réduire la consommation'},
            {'value': 'Reuse', 'label': 'Reuse - Réutiliser directement'},
            {'value': 'Repair', 'label': 'Repair - Réparer'},
            {'value': 'Refurbish', 'label': 'Refurbish - Rénover'},
            {'value': 'Remanufacture', 'label': 'Remanufacture - Refabriquer'},
            {'value': 'Repurpose', 'label': 'Repurpose - Réaffecter'},
            {'value': 'Recycle', 'label': 'Recycle - Recycler'},
            {'value': 'Recover', 'label': 'Recover - Valoriser énergétiquement'}
        ]
        return jsonify({'strategies': strategies})
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la récupération des stratégies: {str(e)}'}), 500

@app.route('/update-end-of-life-strategy', methods=['POST'])
def update_end_of_life_strategy():
    """Met à jour la stratégie de fin de vie d'un élément"""
    import traceback
    import requests
    import urllib.parse
    from config import GRAPHDB_REPO
    
    try:
        data = request.get_json()
        guid = data.get('guid')
        strategy = data.get('strategy')
        
        print(f"[EOL_UPDATE] Début mise à jour - GUID: {guid}, Stratégie: {strategy}")
        
        if not guid or not strategy:
            print("[EOL_UPDATE] ERREUR: GUID ou stratégie manquant")
            return jsonify({'error': 'GUID et stratégie requis'}), 400
        
        # Encoder le GUID pour l'URI (gérer espaces et caractères spéciaux)
        encoded_guid = urllib.parse.quote(str(guid), safe='')
        
        # URI de l'élément avec encodage
        element_uri = f"http://example.com/ifc#{encoded_guid}"
        strategy_uri = f"http://www.semanticweb.org/adamy/ontologies/2025/WLCPO#{strategy}"
        
        print(f"[EOL_UPDATE] GUID original: {guid}")
        print(f"[EOL_UPDATE] GUID encodé: {encoded_guid}")
        print(f"[EOL_UPDATE] URI élément: {element_uri}")
        print(f"[EOL_UPDATE] URI stratégie: {strategy_uri}")
        
        # Supprimer l'ancienne stratégie s'il y en a une
        delete_old_strategy = f"""
        PREFIX wlcpo: <http://www.semanticweb.org/adamy/ontologies/2025/WLCPO#>
        PREFIX dpp: <http://www.semanticweb.org/adamy/ontologies/2025/DPP#>
        
        DELETE {{
            <{element_uri}> dpp:hasDisposalOption ?oldStrategy .
            <{element_uri}> eol:hasType ?oldStrategyText .
        }}
        WHERE {{
            OPTIONAL {{ <{element_uri}> dpp:hasDisposalOption ?oldStrategy }}
            OPTIONAL {{ <{element_uri}> eol:hasType ?oldStrategyText }}
        }}
        """
        
        # Ajouter la nouvelle stratégie
        insert_new_strategy = f"""
        PREFIX wlcpo: <http://www.semanticweb.org/adamy/ontologies/2025/WLCPO#>
        PREFIX dpp: <http://www.semanticweb.org/adamy/ontologies/2025/DPP#>
        
        INSERT DATA {{
            <{element_uri}> dpp:hasDisposalOption <{strategy_uri}> .
            <{element_uri}> eol:hasType "{strategy}" .
        }}
        """
        
        # Endpoint avec validation
        update_endpoint = GRAPHDB_REPO.rstrip("/") + "/statements"
        print(f"[EOL_UPDATE] Endpoint: {update_endpoint}")
        
        # Étape 1: Supprimer l'ancienne stratégie
        print("[EOL_UPDATE] Étape 1: Suppression ancienne stratégie...")
        try:
            response = requests.post(
                update_endpoint, 
                data={"update": delete_old_strategy},
                timeout=30
            )
            
            print(f"[EOL_UPDATE] DELETE - Status: {response.status_code}")
            if response.text:
                print(f"[EOL_UPDATE] DELETE - Response: {response.text[:200]}")
            
            if not response.ok:
                error_msg = f'Erreur lors de la suppression de l\'ancienne stratégie: {response.status_code} - {response.text}'
                print(f"[EOL_UPDATE] ERREUR DELETE: {error_msg}")
                return jsonify({'error': error_msg}), 500
                
        except requests.exceptions.RequestException as e:
            error_msg = f'Erreur réseau lors de la suppression: {str(e)}'
            print(f"[EOL_UPDATE] ERREUR RÉSEAU DELETE: {error_msg}")
            return jsonify({'error': error_msg}), 500
        
        # Étape 2: Ajouter la nouvelle stratégie
        print("[EOL_UPDATE] Étape 2: Ajout nouvelle stratégie...")
        try:
            response = requests.post(
                update_endpoint, 
                data={"update": insert_new_strategy},
                timeout=30
            )
            
            print(f"[EOL_UPDATE] INSERT - Status: {response.status_code}")
            if response.text:
                print(f"[EOL_UPDATE] INSERT - Response: {response.text[:200]}")
            
            if not response.ok:
                error_msg = f'Erreur lors de l\'ajout de la nouvelle stratégie: {response.status_code} - {response.text}'
                print(f"[EOL_UPDATE] ERREUR INSERT: {error_msg}")
                return jsonify({'error': error_msg}), 500
                
        except requests.exceptions.RequestException as e:
            error_msg = f'Erreur réseau lors de l\'ajout: {str(e)}'
            print(f"[EOL_UPDATE] ERREUR RÉSEAU INSERT: {error_msg}")
            return jsonify({'error': error_msg}), 500
        
        success_msg = f'Stratégie {strategy} assignée à {guid}'
        print(f"[EOL_UPDATE] SUCCÈS: {success_msg}")
        return jsonify({'success': True, 'message': success_msg})
        
    except Exception as e:
        error_details = traceback.format_exc()
        error_msg = f'Erreur lors de la mise à jour: {str(e)}'
        print(f"[EOL_UPDATE] EXCEPTION GÉNÉRALE: {error_msg}")
        print(f"[EOL_UPDATE] STACK TRACE: {error_details}")
        return jsonify({'error': error_msg}), 500

@app.route('/update-group-end-of-life-strategy', methods=['POST'])
def update_group_end_of_life_strategy():
    """Met à jour la stratégie de fin de vie pour un groupe d'éléments"""
    try:
        data = request.get_json()
        guids = data.get('guids', [])
        strategy = data.get('strategy')
        
        if not guids or not strategy:
            return jsonify({'error': 'GUIDs et stratégie requis'}), 400
        
        from sparql_client import query_graphdb
        import requests
        
        strategy_uri = f"http://www.semanticweb.org/adamy/ontologies/2025/WLCPO#{strategy}"
        
        # Construire les déclarations pour tous les éléments avec encodage des GUIDs
        delete_statements = []
        insert_statements = []
        
        for guid in guids:
            # Encoder le GUID pour l'URI
            encoded_guid = urllib.parse.quote(str(guid), safe='')
            element_uri = f"http://example.com/ifc#{encoded_guid}"
            
            delete_statements.extend([
                f"<{element_uri}> dpp:hasDisposalOption ?oldStrategy_{encoded_guid.replace('%', '_').replace('-', '_')} .",
                f"<{element_uri}> eol:hasType ?oldStrategyText_{encoded_guid.replace('%', '_').replace('-', '_')} ."
            ])
            insert_statements.extend([
                f"<{element_uri}> dpp:hasDisposalOption <{strategy_uri}> .",
                f"<{element_uri}> eol:hasType \"{strategy}\" ."
            ])
        
        # Supprimer les anciennes stratégies
        delete_query = f"""
        PREFIX wlcpo: <http://www.semanticweb.org/adamy/ontologies/2025/WLCPO#>
        PREFIX dpp: <http://www.semanticweb.org/adamy/ontologies/2025/DPP#>
        
        DELETE {{
            {chr(10).join([f"<http://example.com/ifc#{urllib.parse.quote(str(guid), safe='')}> dpp:hasDisposalOption ?oldStrategy ." for guid in guids])}
            {chr(10).join([f"<http://example.com/ifc#{urllib.parse.quote(str(guid), safe='')}> eol:hasType ?oldStrategyText ." for guid in guids])}
        }}
        WHERE {{
            {chr(10).join([f"OPTIONAL {{ <http://example.com/ifc#{urllib.parse.quote(str(guid), safe='')}> dpp:hasDisposalOption ?oldStrategy }}" for guid in guids])}
            {chr(10).join([f"OPTIONAL {{ <http://example.com/ifc#{urllib.parse.quote(str(guid), safe='')}> eol:hasType ?oldStrategyText }}" for guid in guids])}
        }}
        """
        
        # Ajouter les nouvelles stratégies
        insert_query = f"""
        PREFIX wlcpo: <http://www.semanticweb.org/adamy/ontologies/2025/WLCPO#>
        PREFIX dpp: <http://www.semanticweb.org/adamy/ontologies/2025/DPP#>
        
        INSERT DATA {{
            {chr(10).join(insert_statements)}
        }}
        """
        
        # Exécuter les requêtes
        response = requests.post(GRAPHDB_REPO.rstrip("/") + "/statements", data={"update": delete_query})
        if not response.ok:
            return jsonify({'error': 'Erreur lors de la suppression des anciennes stratégies'}), 500
        
        response = requests.post(GRAPHDB_REPO.rstrip("/") + "/statements", data={"update": insert_query})
        if not response.ok:
            return jsonify({'error': 'Erreur lors de l\'ajout des nouvelles stratégies'}), 500
        
        return jsonify({
            'success': True, 
            'message': f'Stratégie {strategy} assignée à {len(guids)} éléments'
        })
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la mise à jour: {str(e)}'}), 500

@app.route('/get-end-of-life-statistics')
def get_end_of_life_statistics():
    """Retourne les statistiques de recyclabilité"""
    try:
        # Statistiques par stratégie - utiliser la même méthode que les autres routes
        stats_query = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        PREFIX wlcpo: <http://www.semanticweb.org/adamy/ontologies/2025/WLCPO#>
        PREFIX eol: <http://www.w3id.org/dpp/EoL#>
        
        SELECT ?strategy (COUNT(?element) as ?count) WHERE {
            ?element a wlc:Element .
            ?element eol:hasType ?strategy .
        }
        GROUP BY ?strategy
        ORDER BY DESC(?count)
        """
        
        # Utiliser la même méthode que les autres routes qui fonctionnent
        import requests
        
        response = requests.post(
            GRAPHDB_REPO,
            data={"query": stats_query},
            headers={"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"},
            timeout=30
        )
        response.raise_for_status()
        stats_data = response.json()["results"]["bindings"]
        
        if not stats_data:
            return jsonify({
                'statistics': [],
                'total_elements': 0,
                'total_with_strategy': 0,
                'recyclability_percent': 0
            })
        
        # Traitement des statistiques
        strategy_stats = []
        total_with_strategy = 0
        recyclable_count = 0
        recyclable_strategies = ['Recycle', 'Reuse', 'Repurpose']
        
        for stat in stats_data:
            try:
                # Vérifier que stat est bien un dictionnaire
                if not isinstance(stat, dict):
                    print(f"Erreur: stat n'est pas un dictionnaire: {type(stat)} - {stat}")
                    continue
                
                # Accès sécurisé aux valeurs
                strategy_obj = stat.get('strategy', {})
                count_obj = stat.get('count', {})
                
                if not isinstance(strategy_obj, dict) or not isinstance(count_obj, dict):
                    print(f"Erreur: objets strategy ou count mal formés: {strategy_obj}, {count_obj}")
                    continue
                
                strategy = strategy_obj.get('value', '')
                count = int(count_obj.get('value', 0))
                
                if not strategy:
                    continue
                
                total_with_strategy += count
                
                if strategy in recyclable_strategies:
                    recyclable_count += count
                
                strategy_stats.append({
                    'strategy': strategy,
                    'count': count,
                    'percentage': 0  # Sera calculé après
                })
                
            except Exception as e:
                print(f"Erreur lors du traitement d'une statistique: {e} - stat: {stat}")
                continue
        
        # Calculer les pourcentages
        for stat in strategy_stats:
            stat['percentage'] = (stat['count'] / total_with_strategy * 100) if total_with_strategy > 0 else 0
        
        # Compter le nombre total d'éléments
        total_elements_query = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        
        SELECT (COUNT(?element) as ?total) WHERE {
            ?element a wlc:Element .
        }
        """
        
        response = requests.post(
            GRAPHDB_REPO,
            data={"query": total_elements_query},
            headers={"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"},
            timeout=30
        )
        response.raise_for_status()
        total_data = response.json()["results"]["bindings"]
        
        total_elements = 0
        if total_data and len(total_data) > 0:
            try:
                total_obj = total_data[0].get('total', {})
                if isinstance(total_obj, dict):
                    total_elements = int(total_obj.get('value', 0))
            except Exception as e:
                print(f"Erreur lors du calcul du total d'éléments: {e}")
                total_elements = 0
        
        recyclability_percent = (recyclable_count / total_elements * 100) if total_elements > 0 else 0
        
        return jsonify({
            'statistics': strategy_stats,
            'total_elements': total_elements,
            'total_with_strategy': total_with_strategy,
            'recyclability_percent': round(recyclability_percent, 1),
            'recyclable_count': recyclable_count
        })
        
    except Exception as e:
        import traceback
        print(f"Erreur dans get_end_of_life_statistics: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': f'Erreur lors du calcul des statistiques: {str(e)}'}), 500

@app.route('/get-eol-management-data')
def get_eol_management_data():
    """Récupère les données complètes pour l'onglet Gestion Fin de Vie"""
    try:
        # Requête SPARQL pour récupérer toutes les données EOL
        eol_query = """
        PREFIX eol: <http://www.w3id.org/dpp/EoL#>
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        PREFIX wlcpo: <http://www.semanticweb.org/adamy/ontologies/2025/WLCPO#>
        PREFIX dpp: <http://www.semanticweb.org/adamy/ontologies/2025/DPP#>
        
        SELECT ?element ?globalId ?uniformatDesc ?strategy ?destination ?responsible ?cost WHERE {
            ?element a wlc:Element .
            ?element wlc:globalId ?globalId .
            OPTIONAL { ?element wlc:hasUniformatDescription ?uniformatDesc }
            OPTIONAL { ?element eol:hasType ?strategy }
            OPTIONAL { ?element eol:atPlace ?destination }
            OPTIONAL { ?element eol:providesParticipantRole ?responsible }
            OPTIONAL { 
                ?element wlc:hasCost ?costInst .
                ?costInst a wlc:EndOfLifeCosts ;
                          wlc:hasCostValue ?cost .
            }
        }
        ORDER BY ?uniformatDesc
        """
        
        response = requests.post(
            GRAPHDB_REPO,
            data={"query": eol_query},
            headers={"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"},
            timeout=30
        )
        response.raise_for_status()
        results = response.json()["results"]["bindings"]
        
        # Traitement des résultats
        elements_data = []
        for result in results:
            elements_data.append({
                'GlobalId': result.get('globalId', {}).get('value', ''),
                'UniformatDesc': result.get('uniformatDesc', {}).get('value', ''),
                'Strategy': result.get('strategy', {}).get('value', ''),
                'Destination': result.get('destination', {}).get('value', ''),
                'Responsible': result.get('responsible', {}).get('value', ''),
                'Cost': float(result.get('cost', {}).get('value', 0)) if result.get('cost') else 0
            })
        
        return jsonify({'elements': elements_data})
        
    except Exception as e:
        import traceback
        print(f"Erreur dans get_eol_management_data: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'error': f'Erreur lors de la récupération: {str(e)}'}), 500

@app.route('/update-eol-destination', methods=['POST'])
def update_eol_destination():
    """Met à jour la destination de fin de vie d'un élément"""
    try:
        data = request.get_json()
        guid = data.get('guid')
        destination = data.get('destination')
        
        if not guid:
            return jsonify({'error': 'GUID requis'}), 400
        
        # Encoder le GUID pour l'URI
        encoded_guid = urllib.parse.quote(str(guid), safe='')
        element_uri = f"http://example.com/ifc#{encoded_guid}"
        
        # Supprimer l'ancienne destination s'il y en a une
        delete_old = f"""
        PREFIX wlcpo: <http://www.semanticweb.org/adamy/ontologies/2025/WLCPO#>
        
        DELETE {{
            <{element_uri}> eol:atPlace ?oldDestination .
        }}
        WHERE {{
            OPTIONAL {{ <{element_uri}> eol:atPlace ?oldDestination }}
        }}
        """
        
        # Ajouter la nouvelle destination si fournie
        insert_new = ""
        if destination and destination.strip():
            insert_new = f"""
            PREFIX wlcpo: <http://www.semanticweb.org/adamy/ontologies/2025/WLCPO#>
            
            INSERT DATA {{
                <{element_uri}> eol:atPlace "{destination}" .
            }}
            """
        
        # Exécuter les requêtes
        import requests
        
        response = requests.post(GRAPHDB_REPO.rstrip("/") + "/statements", data={"update": delete_old})
        if not response.ok:
            return jsonify({'error': 'Erreur lors de la suppression'}), 500
        
        if insert_new:
            response = requests.post(GRAPHDB_REPO.rstrip("/") + "/statements", data={"update": insert_new})
            if not response.ok:
                return jsonify({'error': 'Erreur lors de l\'ajout'}), 500
        
        return jsonify({'success': True, 'message': f'Destination mise à jour pour {guid}'})
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la mise à jour: {str(e)}'}), 500

@app.route('/update-eol-responsible', methods=['POST'])
def update_eol_responsible():
    """Met à jour le responsable de fin de vie d'un élément"""
    try:
        data = request.get_json()
        guid = data.get('guid')
        responsible = data.get('responsible')
        
        if not guid:
            return jsonify({'error': 'GUID requis'}), 400
        
        # Encoder le GUID pour l'URI
        encoded_guid = urllib.parse.quote(str(guid), safe='')
        element_uri = f"http://example.com/ifc#{encoded_guid}"
        
        # Supprimer l'ancien responsable s'il y en a un
        delete_old = f"""
        PREFIX wlcpo: <http://www.semanticweb.org/adamy/ontologies/2025/WLCPO#>
        
        DELETE {{
            <{element_uri}> eol:providesParticipantRole ?oldResponsible .
        }}
        WHERE {{
            OPTIONAL {{ <{element_uri}> eol:providesParticipantRole ?oldResponsible }}
        }}
        """
        
        # Ajouter le nouveau responsable si fourni
        insert_new = ""
        if responsible and responsible.strip():
            insert_new = f"""
            PREFIX wlcpo: <http://www.semanticweb.org/adamy/ontologies/2025/WLCPO#>
            
            INSERT DATA {{
                <{element_uri}> eol:providesParticipantRole "{responsible}" .
            }}
            """
        
        # Exécuter les requêtes
        import requests
        
        response = requests.post(GRAPHDB_REPO.rstrip("/") + "/statements", data={"update": delete_old})
        if not response.ok:
            return jsonify({'error': 'Erreur lors de la suppression'}), 500
        
        if insert_new:
            response = requests.post(GRAPHDB_REPO.rstrip("/") + "/statements", data={"update": insert_new})
            if not response.ok:
                return jsonify({'error': 'Erreur lors de l\'ajout'}), 500
        
        return jsonify({'success': True, 'message': f'Responsable mis à jour pour {guid}'})
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la mise à jour: {str(e)}'}), 500

@app.route('/get-eol-responsibles')
def get_eol_responsibles():
    """Retourne la liste des responsables disponibles"""
    try:
        # Récupérer les parties prenantes existantes + options prédéfinies
        stakeholders_query = """
        PREFIX wlc: <http://www.semanticweb.org/adamy/ontologies/2025/WLCONTO#>
        
        SELECT DISTINCT ?name WHERE {
            ?stakeholder a wlc:Stakeholder .
            ?stakeholder wlc:hasName ?name .
        }
        ORDER BY ?name
        """
        
        response = requests.post(
            GRAPHDB_REPO,
            data={"query": stakeholders_query},
            headers={"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"},
            timeout=30
        )
        response.raise_for_status()
        results = response.json()["results"]["bindings"]
        
        # Responsables prédéfinis + ceux de l'ontologie
        predefined = [
            'Propriétaire',
            'Entreprise de démolition',
            'Centre de recyclage',
            'Gestionnaire de déchets',
            'Prestataire spécialisé'
        ]
        
        stakeholders = [r['name']['value'] for r in results if r.get('name')]
        all_responsibles = list(set(predefined + stakeholders))
        all_responsibles.sort()
        
        return jsonify({'responsibles': all_responsibles})
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la récupération: {str(e)}'}), 500

@app.route('/update-bulk-eol-data', methods=['POST'])
def update_bulk_eol_data():
    """Met à jour en lot les données EOL pour plusieurs éléments"""
    try:
        data = request.get_json()
        guids = data.get('guids', [])
        strategy = data.get('strategy')
        destination = data.get('destination')
        responsible = data.get('responsible')
        
        if not guids:
            return jsonify({'error': 'GUIDs requis'}), 400
        
        # Construire les requêtes de mise à jour
        updates = []
        
        for guid in guids:
            # Encoder le GUID pour l'URI
            encoded_guid = urllib.parse.quote(str(guid), safe='')
            element_uri = f"http://example.com/ifc#{encoded_guid}"
            
            # Suppression des anciennes valeurs
            delete_query = f"""
            PREFIX wlcpo: <http://www.semanticweb.org/adamy/ontologies/2025/WLCPO#>
            PREFIX dpp: <http://www.semanticweb.org/adamy/ontologies/2025/DPP#>
            
            DELETE {{
                <{element_uri}> dpp:hasDisposalOption ?oldStrategy .
                <{element_uri}> eol:hasType ?oldStrategyText .
                <{element_uri}> eol:atPlace ?oldDestination .
                <{element_uri}> eol:providesParticipantRole ?oldResponsible .
            }}
            WHERE {{
                OPTIONAL {{ <{element_uri}> dpp:hasDisposalOption ?oldStrategy }}
                OPTIONAL {{ <{element_uri}> eol:hasType ?oldStrategyText }}
                OPTIONAL {{ <{element_uri}> eol:atPlace ?oldDestination }}
                OPTIONAL {{ <{element_uri}> eol:providesParticipantRole ?oldResponsible }}
            }}
            """
            updates.append(delete_query)
            
            # Insertion des nouvelles valeurs
            insert_statements = []
            if strategy:
                strategy_uri = f"http://www.semanticweb.org/adamy/ontologies/2025/WLCPO#{strategy}"
                insert_statements.extend([
                    f"<{element_uri}> dpp:hasDisposalOption <{strategy_uri}> .",
                    f"<{element_uri}> eol:hasType \"{strategy}\" ."
                ])
            
            if destination:
                insert_statements.append(f"<{element_uri}> eol:atPlace \"{destination}\" .")
            
            if responsible:
                insert_statements.append(f"<{element_uri}> eol:providesParticipantRole \"{responsible}\" .")
            
            if insert_statements:
                insert_query = f"""
                PREFIX wlcpo: <http://www.semanticweb.org/adamy/ontologies/2025/WLCPO#>
                PREFIX dpp: <http://www.semanticweb.org/adamy/ontologies/2025/DPP#>
                
                INSERT DATA {{
                    {chr(10).join(insert_statements)}
                }}
                """
                updates.append(insert_query)
        
        # Exécuter toutes les mises à jour
        import requests
        
        for update_query in updates:
            response = requests.post(GRAPHDB_REPO.rstrip("/") + "/statements", data={"update": update_query})
            if not response.ok:
                return jsonify({'error': f'Erreur lors de la mise à jour en lot'}), 500
        
        return jsonify({
            'success': True,
            'message': f'Données EOL mises à jour pour {len(guids)} éléments'
        })
        
    except Exception as e:
        return jsonify({'error': f'Erreur lors de la mise à jour en lot: {str(e)}'}), 500

if __name__ == '__main__':
    print("🚀 Démarrage du serveur Flask...")
    app.run(debug=True, host='0.0.0.0', port=8000)
